#!/usr/bin/env python3
"""Generate Roadmap_Переезд.xlsx — migration tracking workbook.

Sheets: Справочники, DATA, Dashboard (Segment×Instrument matrix),
        5× category dashboards (Product×Instrument), Timeline, 3× view stubs.
"""

import datetime
import quopri
import re
from html.parser import HTMLParser
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Paths ─────────────────────────────────────────────────────────────
VAULT = Path(__file__).resolve().parent.parent.parent
ROADMAP_DIR = VAULT / "Atlas" / "Knowledge" / "Переезд Roadmap"
OUTPUT = ROADMAP_DIR / "Roadmap_Переезд.xlsx"

# ── Constants ─────────────────────────────────────────────────────────
CATEGORIES = ["PRE", "1", "2", "3", "4"]
STAGES = [("ИТ", 0.10), ("1%", 0.20), ("5%", 0.40), ("50%", 0.75), ("100%", 1.00)]
SUBSEGMENTS = [("Обычная (все грейсы)", [("до 200к", 0.5), ("свыше 200к", 0.5)])]

# Column indices (0-based) for DATA sheet — grouped logically
class C:
    # Group 0: ID (A-I) — always visible, frozen
    AGG=0; PROD=1; SUBPROD=2; SUBSEG=3; CAT=4; SEG=5; IGRP=6; INSTR=7; ACTIVE=8
    # Group 1: Plan dates (J-N)
    IT_P=9; P1_P=10; P5_P=11; P50_P=12; P100_P=13
    # Group 2: Fact dates (O-S)
    IT_F=14; P1_F=15; P5_F=16; P50_F=17; P100_F=18
    # Group 3: Status (T-V) — formulas
    CUR=19; PROG=20; RAG=21
    # Group 4: Analysis (W-AA) — formulas, collapsed
    NEXT=22; SLIP=23; WABS=24; WSTAT=25; DATOK=26
    # Group 5: Baseline (AB-AF) — hidden
    IT_BL=27; P1_BL=28; P5_BL=29; P50_BL=30; P100_BL=31
    # Group 6: Notes (AG-AI) — collapsed
    EPICS=32; COMMENT=33; UPDATED=34
    TOTAL = 35

def cl(i):
    """0-based index → Excel column letter."""
    return get_column_letter(i + 1)

# Colors
BLUE_H = "D6E4F0"; GREEN_H = "E2EFDA"; GRAY_H = "F2F2F2"; YELLOW_H = "FFF2CC"
RED_F = "FFC7CE"; AMBER_F = "FFEB9C"; GREEN_F = "C6EFCE"; DONE_F = "BDD7EE"
HDR_BG = "2F5496"

THIN = Border(*(Side(style="thin", color="D9D9D9") for _ in range(4)))

def _hdr_style():
    return dict(font=Font(bold=True, size=10, color="FFFFFF"),
                fill=PatternFill(start_color=HDR_BG, end_color=HDR_BG, fill_type="solid"),
                alignment=Alignment(horizontal="center", wrap_text=True))

def _fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

# ── DATA column spec ──────────────────────────────────────────────────
# (header, width, fill_color, is_formula, number_format)
DATA_COLS = [
    # Group 0: ID (A-I)
    ("Агрегация",           16, None,     False, None),
    ("Продукт",             16, None,     False, None),
    ("Подпродукт",          22, None,     False, None),
    ("Подсегмент",          14, None,     False, None),
    ("Категория ПЗ",        12, None,     False, None),
    ("Сегмент",             28, GRAY_H,   True,  None),
    ("Группа инструмента",  20, None,     False, None),
    ("Инструмент",          22, None,     False, None),
    ("Активен",              9, None,     False, None),
    # Group 1: Plan dates (J-N)
    ("ИТ план",             11, BLUE_H,   False, 'DD.MM.YY'),
    ("1% план",             11, BLUE_H,   False, 'DD.MM.YY'),
    ("5% план",             11, BLUE_H,   False, 'DD.MM.YY'),
    ("50% план",            11, BLUE_H,   False, 'DD.MM.YY'),
    ("100% план",           11, BLUE_H,   False, 'DD.MM.YY'),
    # Group 2: Fact dates (O-S)
    ("ИТ факт",             11, GREEN_H,  False, 'DD.MM.YY'),
    ("1% факт",             11, GREEN_H,  False, 'DD.MM.YY'),
    ("5% факт",             11, GREEN_H,  False, 'DD.MM.YY'),
    ("50% факт",            11, GREEN_H,  False, 'DD.MM.YY'),
    ("100% факт",           11, GREEN_H,  False, 'DD.MM.YY'),
    # Group 3: Status (T-V)
    ("Текущий этап",        13, GRAY_H,   True,  None),
    ("Прогресс %",          11, GRAY_H,   True,  '0%'),
    ("RAG",                  7, GRAY_H,   True,  None),
    # Group 4: Analysis (W-AA) — collapsed
    ("Следующий план",      14, GRAY_H,   True,  'DD.MM.YY'),
    ("Сдвиг дней",          11, GRAY_H,   True,  '0'),
    ("Вес абс",              9, GRAY_H,   True,  '0.00%'),
    ("Статус веса",         10, GRAY_H,   True,  None),
    ("Порядок дат ОК",      13, GRAY_H,   True,  None),
    # Group 5: Baseline (AB-AF) — hidden
    ("ИТ baseline",         11, YELLOW_H, False, 'DD.MM.YY'),
    ("1% baseline",         11, YELLOW_H, False, 'DD.MM.YY'),
    ("5% baseline",         11, YELLOW_H, False, 'DD.MM.YY'),
    ("50% baseline",        11, YELLOW_H, False, 'DD.MM.YY'),
    ("100% baseline",       11, YELLOW_H, False, 'DD.MM.YY'),
    # Group 6: Notes (AG-AI) — collapsed
    ("Эпики",               30, None,     False, None),
    ("Комментарии",         25, None,     False, None),
    ("Обновлено",           11, None,     False, 'DD.MM.YY'),
]

# ── HTML Parser ───────────────────────────────────────────────────────

class ConfluenceTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.in_table = self.in_row = self.in_cell = False
        self.rows, self._row, self._cell = [], [], ""
        self._col, self._rowspans, self._colspan, self._rowspan = 0, {}, 1, 1

    def handle_starttag(self, tag, attrs):
        a = dict(attrs)
        if tag == "table": self.in_table = True
        elif tag == "tr" and self.in_table:
            self.in_row = True; self._row = []; self._col = 0
        elif tag in ("td", "th") and self.in_row:
            self.in_cell = True; self._cell = ""
            while self._col in self._rowspans and self._rowspans[self._col] > 0:
                self._row.append(self._rowspans.get(f"{self._col}_v", ""))
                self._rowspans[self._col] -= 1
                if self._rowspans[self._col] == 0:
                    del self._rowspans[self._col]; self._rowspans.pop(f"{self._col}_v", None)
                self._col += 1
            self._colspan = int(a.get("colspan", "1")); self._rowspan = int(a.get("rowspan", "1"))

    def handle_data(self, data):
        if self.in_cell: self._cell += data.strip()

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self.in_cell:
            self.in_cell = False; v = self._cell.strip()
            for i in range(self._colspan):
                self._row.append(v if i == 0 else "")
                if self._rowspan > 1:
                    self._rowspans[self._col] = self._rowspan - 1
                    self._rowspans[f"{self._col}_v"] = v
                self._col += 1
        elif tag == "tr" and self.in_row:
            self.in_row = False
            while self._col in self._rowspans and self._rowspans[self._col] > 0:
                self._row.append(self._rowspans.get(f"{self._col}_v", ""))
                self._rowspans[self._col] -= 1
                if self._rowspans[self._col] == 0:
                    del self._rowspans[self._col]; self._rowspans.pop(f"{self._col}_v", None)
                self._col += 1
            self.rows.append(self._row)
        elif tag == "table": self.in_table = False


def _decode_confluence(path):
    raw = path.read_text(encoding="utf-8")
    s, e = raw.find("<html"), raw.find("</html>") + 7
    return quopri.decodestring(raw[s:e].encode()).decode("utf-8")

def _pct(s):
    s = s.strip().replace("%", "").replace(",", ".")
    try: return float(s) / 100.0 if s else 0.0
    except ValueError: return 0.0

def parse_products(path):
    html = _decode_confluence(path)
    p = ConfluenceTableParser(); p.feed(html)
    items = []
    for row in p.rows[1:]:
        while len(row) < 6: row.append("")
        sub = row[4].strip() if row[4].strip() else row[2].strip()
        items.append(dict(agg=row[0].strip(), w_agg=_pct(row[1]),
                          prod=row[2].strip(), w_prod=_pct(row[3]),
                          subprod=sub, w_subprod=_pct(row[5])))
    return items

def _clean_name(s):
    """Strip parenthetical notes from instrument names."""
    return re.sub(r'\s*\(.*$', '', s).strip()

def parse_instruments(path):
    html = _decode_confluence(path)
    body = html[html.find("<body"):html.find("</body>")]
    items = []
    for grp_html, sub_html in re.findall(r'<li><p[^>]*>(.*?)</p>\s*(?:<ol[^>]*>(.*?)</ol>)?', body, re.DOTALL):
        grp = _clean_name(re.sub(r'<[^>]+>', '', grp_html).strip())
        if sub_html.strip():
            subs = re.findall(r'<li>.*?<p[^>]*>(.*?)</p>.*?</li>', sub_html, re.DOTALL)
            if not subs: subs = re.findall(r'<li><span>(.*?)</span></li>', sub_html, re.DOTALL)
            for si in subs:
                items.append(dict(group=grp, instrument=_clean_name(re.sub(r'<[^>]+>', '', si).strip())))
        else:
            items.append(dict(group=grp, instrument=grp))
    return items


# ── Справочники ───────────────────────────────────────────────────────

def create_справочники(wb, products, instruments):
    ws = wb.active; ws.title = "Справочники"; ws.sheet_properties.tabColor = "4472C4"
    hs = _hdr_style()

    # tblProducts A1:H...
    for ci, h in enumerate(["Ключ","Агрегация","Вес агр","Продукт","Вес прод","Подпродукт","Вес подпрод","Статус веса"], 1):
        c = ws.cell(1, ci, h); c.font = hs["font"]; c.fill = hs["fill"]; c.alignment = hs["alignment"]
    for ri, p in enumerate(products, 2):
        key = f'{p["agg"]}|{p["prod"]}|{p["subprod"]}'
        st = "C" if p["w_agg"] > 0 and p["w_prod"] > 0 and p["w_subprod"] > 0 else "?"
        for ci, v in enumerate([key, p["agg"], p["w_agg"], p["prod"], p["w_prod"], p["subprod"], p["w_subprod"], st], 1):
            c = ws.cell(ri, ci, v)
            if ci in (3,5,7): c.number_format = '0%'
    pe = len(products)+1
    t = Table(displayName="tblProducts", ref=f"A1:H{pe}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(t)
    for ci, w in enumerate([30,18,8,18,8,25,10,10], 1): ws.column_dimensions[cl(ci-1)].width = w

    # tblInstruments J1:K...
    for ci, h in enumerate(["Группа","Инструмент"], 10):
        c = ws.cell(1, ci, h); c.font = hs["font"]; c.fill = hs["fill"]; c.alignment = hs["alignment"]
    for ri, inst in enumerate(instruments, 2):
        ws.cell(ri, 10, inst["group"]); ws.cell(ri, 11, inst["instrument"])
    ie = len(instruments)+1
    t = Table(displayName="tblInstruments", ref=f"J1:K{ie}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(t)
    ws.column_dimensions["J"].width = 25; ws.column_dimensions["K"].width = 28

    # tblStages M1:N...
    for ci, h in enumerate(["Этап","Вес прогресса"], 13):
        c = ws.cell(1, ci, h); c.font = hs["font"]; c.fill = hs["fill"]; c.alignment = hs["alignment"]
    for ri, (s, w) in enumerate(STAGES, 2):
        ws.cell(ri, 13, s); ws.cell(ri, 14, w).number_format = '0%'
    t = Table(displayName="tblStages", ref=f"M1:N{len(STAGES)+1}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(t)

    # tblSubSegments P1:S...
    for ci, h in enumerate(["Ключ","Подпродукт","Подсегмент","Вес"], 16):
        c = ws.cell(1, ci, h); c.font = hs["font"]; c.fill = hs["fill"]; c.alignment = hs["alignment"]
    ri = 2
    for sp, segs in SUBSEGMENTS:
        for sn, sw in segs:
            ws.cell(ri,16,f"{sp}|{sn}"); ws.cell(ri,17,sp); ws.cell(ri,18,sn); ws.cell(ri,19,sw).number_format='0%'
            ri += 1
    se = max(ri-1, 2)
    t = Table(displayName="tblSubSegments", ref=f"P1:S{se}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(t)


# ── DATA ──────────────────────────────────────────────────────────────

def _gen_rows(products, instruments):
    sm = {sp: segs for sp, segs in SUBSEGMENTS}
    rows = []
    for p in products:
        for cat in CATEGORIES:
            seg = f'{p["prod"]} | {cat}'
            for inst in instruments:
                if p["subprod"] in sm:
                    for sn, _ in sm[p["subprod"]]:
                        rows.append([p["agg"], p["prod"], p["subprod"], sn, cat, seg,
                                     inst["group"], inst["instrument"], "Нет"])
                else:
                    rows.append([p["agg"], p["prod"], p["subprod"], "", cat, seg,
                                 inst["group"], inst["instrument"], "Нет"])
    return rows

# Formula builders — all use column constants via cl()
def _f_segment(r):
    return f'={cl(C.PROD)}{r}&" | "&{cl(C.CAT)}{r}'

def _f_weight(r):
    k = f'{cl(C.AGG)}{r}&"|"&{cl(C.PROD)}{r}&"|"&{cl(C.SUBPROD)}{r}'
    sk = f'{cl(C.SUBPROD)}{r}&"|"&{cl(C.SUBSEG)}{r}'
    return (f'=IFERROR(XLOOKUP({k},tblProducts[Ключ],tblProducts[Вес агр],0)'
            f'*XLOOKUP({k},tblProducts[Ключ],tblProducts[Вес прод],0)'
            f'*XLOOKUP({k},tblProducts[Ключ],tblProducts[Вес подпрод],0)'
            f'*IF({cl(C.SUBSEG)}{r}="",1,IFERROR(XLOOKUP({sk},tblSubSegments[Ключ],tblSubSegments[Вес]),1)),0)')

def _f_wstatus(r):
    k = f'{cl(C.AGG)}{r}&"|"&{cl(C.PROD)}{r}&"|"&{cl(C.SUBPROD)}{r}'
    return f'=IFERROR(XLOOKUP({k},tblProducts[Ключ],tblProducts[Статус веса],"?"),"?")'

def _f_stage(r):
    u,s,q,o,m = cl(C.P100_F), cl(C.P50_F), cl(C.P5_F), cl(C.P1_F), cl(C.IT_F)
    return f'=IF({u}{r}<>"","100%",IF({s}{r}<>"","50%",IF({q}{r}<>"","5%",IF({o}{r}<>"","1%",IF({m}{r}<>"","ИТ","Не начат")))))'

def _f_progress(r):
    return f'=IFERROR(XLOOKUP({cl(C.CUR)}{r},tblStages[Этап],tblStages[Вес прогресса],0),0)'

def _f_next(r):
    m,o,q,s,u = cl(C.IT_F),cl(C.P1_F),cl(C.P5_F),cl(C.P50_F),cl(C.P100_F)
    k,n,p,rr,t = cl(C.IT_P),cl(C.P1_P),cl(C.P5_P),cl(C.P50_P),cl(C.P100_P)
    return f'=IF({m}{r}="",{k}{r},IF({o}{r}="",{n}{r},IF({q}{r}="",{p}{r},IF({s}{r}="",{rr}{r},IF({u}{r}="",{t}{r},"")))))'

def _f_slip(r):
    m,o,q,s,u = cl(C.IT_F),cl(C.P1_F),cl(C.P5_F),cl(C.P50_F),cl(C.P100_F)
    k,n,p,rr,t = cl(C.IT_P),cl(C.P1_P),cl(C.P5_P),cl(C.P50_P),cl(C.P100_P)
    v,w,x,y,z = cl(C.IT_BL),cl(C.P1_BL),cl(C.P5_BL),cl(C.P50_BL),cl(C.P100_BL)
    nx = cl(C.NEXT)
    return (f'=IF({nx}{r}=""," ",'
            f'IF({m}{r}="",IF(AND({k}{r}<>"",{v}{r}<>""),{k}{r}-{v}{r},""),'
            f'IF({o}{r}="",IF(AND({n}{r}<>"",{w}{r}<>""),{n}{r}-{w}{r},""),'
            f'IF({q}{r}="",IF(AND({p}{r}<>"",{x}{r}<>""),{p}{r}-{x}{r},""),'
            f'IF({s}{r}="",IF(AND({rr}{r}<>"",{y}{r}<>""),{rr}{r}-{y}{r},""),'
            f'IF({u}{r}="",IF(AND({t}{r}<>"",{z}{r}<>""),{t}{r}-{z}{r},""),""))))))')

def _f_rag(r):
    u, ad = cl(C.P100_F), cl(C.SLIP)
    return f'=IF({u}{r}<>"","DONE",IF(OR({ad}{r}="",{ad}{r}=" "),"—",IF({ad}{r}>14,"RED",IF({ad}{r}>0,"AMBER","GREEN"))))'

def _f_datok(r):
    m,o,q,s,u = cl(C.IT_F),cl(C.P1_F),cl(C.P5_F),cl(C.P50_F),cl(C.P100_F)
    return (f'=AND(IF(AND({m}{r}<>"",{o}{r}<>""),{m}{r}<={o}{r},TRUE),'
            f'IF(AND({o}{r}<>"",{q}{r}<>""),{o}{r}<={q}{r},TRUE),'
            f'IF(AND({q}{r}<>"",{s}{r}<>""),{q}{r}<={s}{r},TRUE),'
            f'IF(AND({s}{r}<>"",{u}{r}<>""),{s}{r}<={u}{r},TRUE))')

FORMULA_MAP = {
    C.SEG: _f_segment, C.WABS: _f_weight, C.WSTAT: _f_wstatus,
    C.CUR: _f_stage, C.PROG: _f_progress, C.NEXT: _f_next,
    C.SLIP: _f_slip, C.RAG: _f_rag, C.DATOK: _f_datok,
}


def create_data(wb, products, instruments):
    ws = wb.create_sheet("DATA"); ws.sheet_properties.tabColor = "00B050"
    hs = _hdr_style()

    # Headers
    for ci, (hdr, w, fill_c, _, _) in enumerate(DATA_COLS):
        c = ws.cell(1, ci+1, hdr)
        c.font = hs["font"]; c.fill = hs["fill"]; c.alignment = hs["alignment"]; c.border = THIN
        ws.column_dimensions[cl(ci)].width = w

    # Data
    data_rows = _gen_rows(products, instruments)
    N = len(data_rows)
    print(f"  Generating {N} DATA rows...")

    for ri_off, rd in enumerate(data_rows):
        r = ri_off + 2
        # Static columns (A-I minus segment)
        for ci in range(min(len(rd), C.ACTIVE+1)):
            if ci == C.SEG: continue  # formula
            ws.cell(r, ci+1, rd[ci]).border = THIN

        # Formula columns
        for col_idx, maker in FORMULA_MAP.items():
            ws.cell(r, col_idx+1, maker(r)).border = THIN

        # Formatting
        for ci, (_, _, fill_c, _, nf) in enumerate(DATA_COLS):
            cell = ws.cell(r, ci+1)
            if nf: cell.number_format = nf
            if fill_c: cell.fill = _fill(fill_c)

    lr = N + 1

    # Table
    t = Table(displayName="tblData", ref=f"A1:{cl(C.TOTAL-1)}{lr}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
    ws.add_table(t)

    ws.freeze_panes = f"{cl(C.IT_P)}2"  # freeze after ID group

    # Column groups (outline collapse/expand)
    ws.column_dimensions.group(cl(C.NEXT), cl(C.DATOK), hidden=True)     # Analysis — collapsed
    ws.column_dimensions.group(cl(C.IT_BL), cl(C.P100_BL), hidden=True)  # Baseline — hidden
    ws.column_dimensions.group(cl(C.EPICS), cl(C.UPDATED), hidden=True)  # Notes — collapsed

    # Validation
    dv = DataValidation(type="list", formula1='"PRE,1,2,3,4"', allow_blank=True, errorStyle="stop")
    dv.add(f"{cl(C.CAT)}2:{cl(C.CAT)}{lr}"); ws.add_data_validation(dv)
    dv2 = DataValidation(type="list", formula1='"Да,Нет"', allow_blank=False, errorStyle="stop")
    dv2.add(f"{cl(C.ACTIVE)}2:{cl(C.ACTIVE)}{lr}"); ws.add_data_validation(dv2)

    for di in [C.IT_P,C.IT_F,C.P1_P,C.P1_F,C.P5_P,C.P5_F,C.P50_P,C.P50_F,C.P100_P,C.P100_F,
               C.IT_BL,C.P1_BL,C.P5_BL,C.P50_BL,C.P100_BL]:
        d = DataValidation(type="date", operator="between", formula1="2025-01-01", formula2="2028-12-31",
                           allow_blank=True, errorStyle="warning")
        d.add(f"{cl(di)}2:{cl(di)}{lr}"); ws.add_data_validation(d)

    # Conditional formatting: RAG
    rag_rng = f"{cl(C.RAG)}2:{cl(C.RAG)}{lr}"
    for val, color in [("RED",RED_F),("AMBER",AMBER_F),("GREEN",GREEN_F),("DONE",DONE_F)]:
        ws.conditional_formatting.add(rag_rng, CellIsRule(operator="equal", formula=[f'"{val}"'], fill=_fill(color)))

    # Date order violation
    ws.conditional_formatting.add(f"{cl(C.DATOK)}2:{cl(C.DATOK)}{lr}",
        CellIsRule(operator="equal", formula=["FALSE"], fill=_fill(RED_F)))

    # Protection
    unlocked = Protection(locked=False)
    formula_cols = {ci for ci, (_, _, _, is_f, _) in enumerate(DATA_COLS) if is_f}
    for r in range(1, lr+1):
        for ci in range(C.TOTAL):
            cell = ws.cell(r, ci+1)
            cell.protection = Protection(locked=True) if (ci in formula_cols or r == 1) else unlocked
    ws.protection.sheet = True; ws.protection.password = "edit"
    ws.protection.formatColumns = ws.protection.formatRows = False
    ws.protection.sort = ws.protection.autoFilter = False

    return ws, N


# ── Dashboard: Segment × Instrument matrix ────────────────────────────

def create_dashboard(wb, products, instruments, total_rows):
    ws = wb.create_sheet("Dashboard"); ws.sheet_properties.tabColor = "FFC000"
    hs = _hdr_style()

    # KPI strip
    ws.merge_cells("A1:F1")
    ws["A1"] = "Roadmap «Переезд» — Сегмент × Инструмент"
    ws["A1"].font = Font(bold=True, size=16, color=HDR_BG)

    kpi_font = Font(bold=True, size=24, color=HDR_BG)
    kpi_lbl = Font(size=10, color="808080")
    for ci, (lbl, fmt, formula) in enumerate([
        ("Общий прогресс", '0.0%',
         '=IFERROR(SUMPRODUCT((tblData[Активен]="Да")*tblData[Вес абс]*tblData[Прогресс %])/SUMPRODUCT((tblData[Активен]="Да")*tblData[Вес абс]),0)'),
        ("Покрытие весами", '0%',
         '=IFERROR(SUMPRODUCT((tblData[Статус веса]="C")*tblData[Вес абс])/SUMPRODUCT((tblData[Активен]="Да")*tblData[Вес абс]),0)'),
        ("RED", '0', '=COUNTIF(tblData[RAG],"RED")'),
        ("DONE", '0', '=COUNTIF(tblData[RAG],"DONE")'),
    ]):
        col_l = cl(ci * 2)
        ws[f"{col_l}2"] = lbl; ws[f"{col_l}2"].font = kpi_lbl
        ws[f"{col_l}3"] = formula; ws[f"{col_l}3"].font = kpi_font; ws[f"{col_l}3"].number_format = fmt

    # Matrix: Row 5 = header, Row 6+ = segments
    MATRIX_ROW = 5
    unique_instr = list(dict.fromkeys(i["instrument"] for i in instruments))

    # Row headers
    for ci, h in enumerate(["Агрегация", "Продукт", "Кат. ПЗ", "Сегмент"]):
        c = ws.cell(MATRIX_ROW, ci+1, h); c.font = hs["font"]; c.fill = hs["fill"]; c.alignment = hs["alignment"]
    ws.column_dimensions["A"].width = 16; ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 8; ws.column_dimensions["D"].width = 26

    # Instrument column headers
    INSTR_COL_START = 5  # E
    for ii, instr in enumerate(unique_instr):
        ci = INSTR_COL_START + ii
        c = ws.cell(MATRIX_ROW, ci, instr)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = _fill(HDR_BG)
        c.alignment = Alignment(horizontal="center", text_rotation=90)
        ws.column_dimensions[cl(ci-1)].width = 6

    # Segment rows: one per unique (Продукт × Категория ПЗ)
    seen = set(); seg_rows = []
    for p in products:
        for cat in CATEGORIES:
            key = (p["agg"], p["prod"], cat)
            if key not in seen:
                seen.add(key); seg_rows.append(key)

    for ri_off, (agg, prod, cat) in enumerate(seg_rows):
        r = MATRIX_ROW + 1 + ri_off
        ws.cell(r, 1, agg); ws.cell(r, 2, prod); ws.cell(r, 3, cat)
        ws.cell(r, 4, f"{prod} | {cat}")

        for ii, instr in enumerate(unique_instr):
            ci = INSTR_COL_START + ii
            # Weighted progress for this segment × instrument
            ws.cell(r, ci,
                value=f'=IFERROR(SUMPRODUCT((tblData[Продукт]=$B{r})*(tblData[Категория ПЗ]=$C{r})'
                      f'*(tblData[Инструмент]={cl(ci-1)}${MATRIX_ROW})*(tblData[Активен]="Да")'
                      f'*tblData[Вес абс]*tblData[Прогресс %])'
                      f'/SUMPRODUCT((tblData[Продукт]=$B{r})*(tblData[Категория ПЗ]=$C{r})'
                      f'*(tblData[Инструмент]={cl(ci-1)}${MATRIX_ROW})*(tblData[Активен]="Да")'
                      f'*tblData[Вес абс]),0)')
            ws.cell(r, ci).number_format = '0%'
            ws.cell(r, ci).alignment = Alignment(horizontal="center")

    last_r = MATRIX_ROW + len(seg_rows)
    last_c = INSTR_COL_START + len(unique_instr) - 1

    # Data bars on matrix cells
    matrix_range = f"{cl(INSTR_COL_START-1)}{MATRIX_ROW+1}:{cl(last_c-1)}{last_r}"
    ws.conditional_formatting.add(matrix_range, DataBarRule(
        start_type="num", start_value=0, end_type="num", end_value=1, color="4472C4"))

    ws.freeze_panes = f"{cl(INSTR_COL_START-1)}{MATRIX_ROW+1}"
    ws.auto_filter.ref = f"A{MATRIX_ROW}:{cl(last_c-1)}{last_r}"
    return ws


# ── Category Dashboards (Product × Instrument per category) ──────────

def create_category_dashboards(wb, products, instruments):
    unique_instr = list(dict.fromkeys(i["instrument"] for i in instruments))
    seen = set(); unique_prods = []
    for p in products:
        key = (p["agg"], p["prod"])
        if key not in seen: seen.add(key); unique_prods.append(key)

    hs = _hdr_style()
    colors = ["C00000", "ED7D31", "FFC000", "70AD47", "4472C4"]

    for cat_idx, cat in enumerate(CATEGORIES):
        ws = wb.create_sheet(f"ПЗ {cat}")
        ws.sheet_properties.tabColor = colors[cat_idx]

        ws["A1"] = f"Категория ПЗ: {cat} — Продукт × Инструмент"
        ws["A1"].font = Font(bold=True, size=14, color=HDR_BG)

        # KPI for this category
        ws["A2"] = "Прогресс:"
        ws["A2"].font = Font(size=10, color="808080")
        ws["B2"] = (f'=IFERROR(SUMPRODUCT((tblData[Активен]="Да")*(tblData[Категория ПЗ]="{cat}")'
                    f'*tblData[Вес абс]*tblData[Прогресс %])/SUMPRODUCT((tblData[Активен]="Да")'
                    f'*(tblData[Категория ПЗ]="{cat}")*tblData[Вес абс]),0)')
        ws["B2"].font = Font(bold=True, size=14, color=HDR_BG); ws["B2"].number_format = '0.0%'

        HR = 4  # header row
        for ci, h in enumerate(["Агрегация", "Продукт"]):
            c = ws.cell(HR, ci+1, h); c.font = hs["font"]; c.fill = hs["fill"]; c.alignment = hs["alignment"]
        ws.column_dimensions["A"].width = 16; ws.column_dimensions["B"].width = 18

        ICOL = 3
        for ii, instr in enumerate(unique_instr):
            ci = ICOL + ii
            c = ws.cell(HR, ci, instr)
            c.font = Font(bold=True, size=9, color="FFFFFF")
            c.fill = _fill(HDR_BG)
            c.alignment = Alignment(horizontal="center", text_rotation=90)
            ws.column_dimensions[cl(ci-1)].width = 6

        for ri_off, (agg, prod) in enumerate(unique_prods):
            r = HR + 1 + ri_off
            ws.cell(r, 1, agg); ws.cell(r, 2, prod)

            for ii, instr in enumerate(unique_instr):
                ci = ICOL + ii
                ws.cell(r, ci,
                    value=f'=IFERROR(SUMPRODUCT((tblData[Продукт]=$B{r})'
                          f'*(tblData[Категория ПЗ]="{cat}")*(tblData[Инструмент]={cl(ci-1)}${HR})'
                          f'*(tblData[Активен]="Да")*tblData[Вес абс]*tblData[Прогресс %])'
                          f'/SUMPRODUCT((tblData[Продукт]=$B{r})*(tblData[Категория ПЗ]="{cat}")'
                          f'*(tblData[Инструмент]={cl(ci-1)}${HR})*(tblData[Активен]="Да")'
                          f'*tblData[Вес абс]),0)')
                ws.cell(r, ci).number_format = '0%'
                ws.cell(r, ci).alignment = Alignment(horizontal="center")

        last_r = HR + len(unique_prods)
        last_ci = ICOL + len(unique_instr) - 1
        matrix_rng = f"{cl(ICOL-1)}{HR+1}:{cl(last_ci-1)}{last_r}"
        ws.conditional_formatting.add(matrix_rng, DataBarRule(
            start_type="num", start_value=0, end_type="num", end_value=1, color=colors[cat_idx]))

        ws.freeze_panes = f"{cl(ICOL-1)}{HR+1}"


# ── Timeline (stage-based) ────────────────────────────────────────────

STAGE_NAMES = ["ИТ", "1%", "5%", "50%", "100%"]
PLAN_COLS_TBL = ["ИТ план", "1% план", "5% план", "50% план", "100% план"]
FACT_COLS_TBL = ["ИТ факт", "1% факт", "5% факт", "50% факт", "100% факт"]

def create_timeline(wb, products, instruments):
    ws = wb.create_sheet("Timeline"); ws.sheet_properties.tabColor = "7030A0"
    hs = _hdr_style()

    ws["A1"] = "Timeline — Этапы миграции"
    ws["A1"].font = Font(bold=True, size=14, color=HDR_BG)
    ws.merge_cells("A1:E1")

    # Two-level header: Row 2 = stage group names, Row 3 = column headers
    GR = 2  # group label row
    HR = 3  # header row

    # Fixed identity columns (A-E)
    fixed = [("Агрегация",16),("Продукт",16),("Кат. ПЗ",10),("Сегмент",26),("Инструмент",22)]
    for ci, (h, w) in enumerate(fixed):
        c = ws.cell(HR, ci+1, h); c.font = hs["font"]; c.fill = hs["fill"]; c.alignment = hs["alignment"]
        ws.column_dimensions[cl(ci)].width = w

    # Stage columns: 5 stages × 2 (план/факт) = columns F-O
    SCOL = 5  # 0-based index of first stage column (F)
    stage_colors = ["B4C6E7", "C5E0B4", "FFE699", "F8CBAD", "D6A4D6"]
    for si, sname in enumerate(STAGE_NAMES):
        pc = SCOL + si * 2      # plan column 0-based
        fc = SCOL + si * 2 + 1  # fact column 0-based
        # Group label row: merge plan+fact cells
        ws.merge_cells(start_row=GR, start_column=pc+1, end_row=GR, end_column=fc+1)
        gc = ws.cell(GR, pc+1, sname)
        gc.font = Font(bold=True, size=11, color="FFFFFF")
        gc.fill = _fill(HDR_BG); gc.alignment = Alignment(horizontal="center")
        # Sub-headers
        for sub_ci, sub_lbl in [(pc, "план"), (fc, "факт")]:
            c = ws.cell(HR, sub_ci+1, sub_lbl)
            c.font = Font(bold=True, size=9, color="FFFFFF")
            c.fill = _fill(HDR_BG); c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cl(sub_ci)].width = 11

    # Summary columns (P-R) after 10 stage columns
    SUMCOL = SCOL + 10  # 0-based = 15
    for si, (h, w, nf) in enumerate([("Прогресс %",11,'0%'), ("RAG",7,None), ("Сдвиг",11,'0')]):
        ci = SUMCOL + si
        c = ws.cell(HR, ci+1, h); c.font = hs["font"]; c.fill = hs["fill"]; c.alignment = hs["alignment"]
        ws.column_dimensions[cl(ci)].width = w

    # Segment × Instrument rows
    seen = set(); prods = []
    for p in products:
        k = (p["agg"], p["prod"])
        if k not in seen: seen.add(k); prods.append(k)

    tl_rows = []
    for agg, prod in prods:
        for cat in CATEGORIES:
            for inst in instruments:
                tl_rows.append((agg, prod, cat, f"{prod} | {cat}", inst["instrument"]))

    print(f"  Timeline: {len(tl_rows)} rows (stage-based)")

    for ri_off, (agg, prod, cat, seg, instr) in enumerate(tl_rows):
        r = HR + 1 + ri_off
        ws.cell(r,1,agg); ws.cell(r,2,prod); ws.cell(r,3,cat); ws.cell(r,4,seg); ws.cell(r,5,instr)

        cr = (f'tblData[Продукт],B{r},tblData[Категория ПЗ],C{r},'
              f'tblData[Инструмент],E{r},tblData[Активен],"Да"')

        # Stage date pairs: plan (MINIFS) and fact (MAXIFS) for each stage
        for si in range(5):
            pc = SCOL + si * 2      # plan col 0-based
            fc = SCOL + si * 2 + 1  # fact col 0-based
            plan_tbl = PLAN_COLS_TBL[si]
            fact_tbl = FACT_COLS_TBL[si]
            ws.cell(r, pc+1, f'=IFERROR(MINIFS(tblData[{plan_tbl}],{cr}),"")').number_format = 'DD.MM.YY'
            ws.cell(r, fc+1, f'=IFERROR(MAXIFS(tblData[{fact_tbl}],{cr}),"")').number_format = 'DD.MM.YY'

        # Progress %
        ws.cell(r, SUMCOL+1,
            f'=IFERROR(SUMPRODUCT((tblData[Продукт]=B{r})*(tblData[Категория ПЗ]=C{r})'
            f'*(tblData[Инструмент]=E{r})*(tblData[Активен]="Да")*tblData[Вес абс]*tblData[Прогресс %])'
            f'/SUMPRODUCT((tblData[Продукт]=B{r})*(tblData[Категория ПЗ]=C{r})'
            f'*(tblData[Инструмент]=E{r})*(tblData[Активен]="Да")*tblData[Вес абс]),0)')
        ws.cell(r, SUMCOL+1).number_format = '0%'

        # RAG (aggregate: RED > AMBER > DONE > GREEN)
        cr2 = f'tblData[Продукт],B{r},tblData[Категория ПЗ],C{r},tblData[Инструмент],E{r}'
        ws.cell(r, SUMCOL+2,
            f'=IF(COUNTIFS({cr2},tblData[RAG],"RED")>0,"RED",'
            f'IF(COUNTIFS({cr2},tblData[RAG],"AMBER")>0,"AMBER",'
            f'IF(AND(COUNTIFS({cr2},tblData[RAG],"DONE")>0,'
            f'COUNTIFS({cr2},tblData[RAG],"DONE")=COUNTIFS({cr},1)),"DONE","GREEN")))')

        # Slippage (max across sub-products)
        ws.cell(r, SUMCOL+3,
            f'=IFERROR(MAXIFS(tblData[Сдвиг дней],{cr}),"")')

    last_r = HR + len(tl_rows)
    LAST_COL = SUMCOL + 2  # 0-based index of last column (RAG+1=Сдвиг is SUMCOL+2)

    # Conditional formatting — fact date cells: green when non-empty
    for si in range(5):
        fc = SCOL + si * 2 + 1  # fact col 0-based
        fact_rng = f"{cl(fc)}{HR+1}:{cl(fc)}{last_r}"
        ws.conditional_formatting.add(fact_rng,
            CellIsRule(operator="notEqual", formula=['""'], fill=_fill(GREEN_F)))

    # RAG column conditional formatting
    rag_rng = f"{cl(SUMCOL+1)}{HR+1}:{cl(SUMCOL+1)}{last_r}"
    for val, color in [("RED",RED_F),("AMBER",AMBER_F),("GREEN",GREEN_F),("DONE",DONE_F)]:
        ws.conditional_formatting.add(rag_rng,
            CellIsRule(operator="equal", formula=[f'"{val}"'], fill=_fill(color)))

    # Progress data bar
    prog_rng = f"{cl(SUMCOL)}{HR+1}:{cl(SUMCOL)}{last_r}"
    ws.conditional_formatting.add(prog_rng, DataBarRule(
        start_type="num", start_value=0, end_type="num", end_value=1, color="4472C4"))

    ws.freeze_panes = f"{cl(SCOL)}{HR+1}"
    ws.auto_filter.ref = f"A{HR}:{cl(LAST_COL)}{last_r}"
    return ws


# ── Gantt ─────────────────────────────────────────────────────────────

from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint

# Base date for Gantt day offsets
GANTT_EPOCH = "DATE(2026,1,1)"

def create_gantt(wb, products, instruments):
    ws = wb.create_sheet("Ганг"); ws.sheet_properties.tabColor = "7030A0"

    ws["A1"] = "Диаграмма Ганта — Переезд"
    ws["A1"].font = Font(bold=True, size=16, color=HDR_BG)
    ws.merge_cells("A1:N1")

    # Unique products
    seen = set(); unique_prods = []
    for p in products:
        k = (p["agg"], p["prod"])
        if k not in seen: seen.add(k); unique_prods.append(k)

    n_prods = len(unique_prods)
    colors = ["C00000", "ED7D31", "FFC000", "70AD47", "4472C4"]
    chart_h_rows = max(n_prods + 4, 18)  # chart height in rows

    HELPER_COL = 23  # 0-based = column X (hidden helper tables)
    chart_row = 3

    for cat_idx, cat in enumerate(CATEGORIES):
        # ── Helper table for this category ──
        ht_start_row = 2 + cat_idx * (n_prods + 3)

        # Headers
        for ci, h in enumerate(["Продукт", "Старт", "Финиш", "Посл.факт",
                                 "Gap", "Выполнено", "Осталось"]):
            c = ws.cell(ht_start_row, HELPER_COL + ci + 1, h)
            c.font = Font(bold=True, size=9)

        for ri, (agg, prod) in enumerate(unique_prods):
            r = ht_start_row + 1 + ri
            ws.cell(r, HELPER_COL + 1, prod)  # A: Product name

            cr = (f'tblData[Продукт],{cl(HELPER_COL)}${r},'
                  f'tblData[Категория ПЗ],"{cat}",tblData[Активен],"Да"')

            # B: Start = earliest ИТ план
            ws.cell(r, HELPER_COL + 2,
                f'=IFERROR(MINIFS(tblData[ИТ план],{cr}),0)')
            # C: End = latest 100% план
            ws.cell(r, HELPER_COL + 3,
                f'=IFERROR(MAXIFS(tblData[100% план],{cr}),0)')
            # D: Latest fact across all 5 fact columns
            fact_parts = []
            for ftbl in FACT_COLS_TBL:
                fact_parts.append(f'IFERROR(MAXIFS(tblData[{ftbl}],{cr},tblData[{ftbl}],">0"),0)')
            ws.cell(r, HELPER_COL + 4, f'=MAX({",".join(fact_parts)})')

            gc = cl(HELPER_COL + 1)  # helper col A (Product)
            sc = cl(HELPER_COL + 1 + 1)  # Start col
            ec = cl(HELPER_COL + 1 + 2)  # End col
            fc = cl(HELPER_COL + 1 + 3)  # LastFact col

            # E: Gap = Start - epoch (days)
            ws.cell(r, HELPER_COL + 5,
                f'=IF({sc}{r}>0,{sc}{r}-{GANTT_EPOCH},0)')
            # F: Done = LastFact - Start (if facts exist)
            ws.cell(r, HELPER_COL + 6,
                f'=IF(AND({sc}{r}>0,{fc}{r}>0),{fc}{r}-{sc}{r},0)')
            # G: Remaining = End - Start - Done
            done_c = cl(HELPER_COL + 1 + 5)
            ws.cell(r, HELPER_COL + 7,
                f'=IF(AND({sc}{r}>0,{ec}{r}>0),MAX({ec}{r}-{sc}{r}-{done_c}{r},0),0)')

        data_first = ht_start_row + 1
        data_last = ht_start_row + n_prods

        # ── Chart ──
        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.gapWidth = 50
        chart.width = 28
        chart.height = max(n_prods * 0.7, 10)
        chart.title = f"ПЗ {cat}"

        # Categories = product names
        cats_ref = Reference(ws, min_col=HELPER_COL + 1, min_row=data_first, max_row=data_last)

        # Data: Gap (col E), Done (col F), Remaining (col G) — include header row
        data_ref = Reference(ws, min_col=HELPER_COL + 5, max_col=HELPER_COL + 7,
                             min_row=ht_start_row, max_row=data_last)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        # Series 0: Gap — invisible
        s0 = chart.series[0]
        s0.graphicalProperties.noFill = True
        s0.graphicalProperties.line.noFill = True

        # Series 1: Done — colored by category
        s1 = chart.series[1]
        s1.graphicalProperties.solidFill = colors[cat_idx]

        # Series 2: Remaining — light gray
        s2 = chart.series[2]
        s2.graphicalProperties.solidFill = "D6DCE4"

        # Reverse Y axis so first product is on top
        chart.y_axis.scaling.orientation = "maxMin"
        chart.y_axis.delete = False
        chart.x_axis.title = "Дни от 01.01.2026"
        chart.x_axis.numFmt = '0'
        chart.legend.position = "b"

        ws.add_chart(chart, f"A{chart_row}")
        chart_row += chart_h_rows

    # Hide helper columns
    ws.column_dimensions.group(cl(HELPER_COL), cl(HELPER_COL + 6), hidden=True)

    print(f"  Gantt: 5 charts × {n_prods} products")
    return ws


# ── Инструкция ────────────────────────────────────────────────────────

def create_instructions(wb):
    ws = wb.create_sheet("Инструкция", 0)  # insert as first sheet
    ws.sheet_properties.tabColor = "4472C4"
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 40

    title_font = Font(bold=True, size=16, color=HDR_BG)
    h2_font = Font(bold=True, size=13, color=HDR_BG)
    h3_font = Font(bold=True, size=11, color="333333")
    body = Font(size=11)
    hint = Font(size=10, italic=True, color="808080")
    hs = _hdr_style()

    r = 1
    ws.cell(r, 1, "Инструкция по работе с Roadmap «Переезд»").font = title_font; r += 2

    # ── Section 1: Overview ──
    ws.cell(r, 1, "1. Что отслеживает этот роадмап").font = h2_font; r += 1
    for line in [
        "Роадмап отслеживает миграцию с текущей архитектуры на целевую.",
        "Три измерения: Продукт (иерархия: Агрегация → Продукт → Подпродукт),",
        "Категория ПЗ (PRE, 1, 2, 3, 4), Инструмент (14 инструментов в 7 группах).",
        "Каждая комбинация проходит 5 этапов: ИТ → 1% → 5% → 50% → 100%.",
        "Прогресс взвешен по весам продуктов из справочника.",
    ]:
        ws.cell(r, 1, line).font = body; r += 1
    r += 1

    # ── Section 2: Sheet guide ──
    ws.cell(r, 1, "2. Карта листов").font = h2_font; r += 1
    for ci, h in enumerate(["Лист", "Назначение", "Когда использовать"]):
        c = ws.cell(r, ci+1, h); c.font = hs["font"]; c.fill = hs["fill"]; c.alignment = hs["alignment"]
    r += 1
    guide = [
        ("DATA", "Все комбинации, плановые и фактические даты", "Ввод данных: активация комбинаций, ввод дат"),
        ("Dashboard", "Матрица Сегмент × Инструмент, KPI", "Общая картина прогресса, статус-митинг"),
        ("ПЗ PRE / 1 / 2 / 3 / 4", "Продукт × Инструмент по категории", "Детальный прогресс по категории ПЗ"),
        ("Timeline", "Этапы миграции с план/факт датами", "Контроль сроков, выявление задержек"),
        ("Ганг", "Диаграмма Ганта по продуктам", "Визуальный обзор таймлайна"),
        ("Справочники", "Продукты, инструменты, этапы, подсегменты", "Проверка/обновление справочных данных"),
    ]
    for name, purpose, when in guide:
        ws.cell(r, 1, name).font = Font(bold=True, size=11)
        ws.cell(r, 2, purpose).font = body
        ws.cell(r, 3, when).font = body
        r += 1
    r += 1

    # ── Section 3: Workflow ──
    ws.cell(r, 1, "3. Рабочий процесс").font = h2_font; r += 1
    steps = [
        ("Шаг 1 — Активация комбинаций",
         'Перейдите на лист DATA. Отфильтруйте по нужному продукту (столбец B). '
         'В столбце I (Активен) поставьте "Да" для актуальных комбинаций подпродукт × инструмент × категория.'),
        ("Шаг 2 — Плановые даты",
         'В столбцах J-N (голубые) введите плановые даты для каждого этапа: ИТ, 1%, 5%, 50%, 100%. '
         'Даты в формате ДД.ММ.ГГ.'),
        ("Шаг 3 — Фактические даты",
         'По мере прогресса заполняйте столбцы O-S (зелёные) — фактические даты завершения этапов.'),
        ("Шаг 4 — Контроль прогресса",
         'Смотрите листы Dashboard, Timeline и Ганг — они обновляются автоматически. '
         'RAG-статус: GREEN = в плане, AMBER = сдвиг 1-14 дней, RED = сдвиг >14 дней, DONE = завершено.'),
        ("Шаг 5 — Перепланирование",
         'При изменении сроков: разверните скрытую группу Baseline (столбцы AB-AF), '
         'скопируйте текущие планы в Baseline, затем обновите плановые даты. '
         'Столбец «Сдвиг дней» покажет разницу с базелайном.'),
    ]
    for title, desc in steps:
        ws.cell(r, 1, title).font = h3_font; r += 1
        ws.cell(r, 1, desc).font = body
        ws.cell(r, 1).alignment = Alignment(wrap_text=True)
        r += 2

    # ── Section 4: Column groups ──
    ws.cell(r, 1, "4. Группы столбцов в DATA").font = h2_font; r += 1
    ws.cell(r, 1, "DATA содержит 35 столбцов, организованных в 7 логических групп. "
                   "Свёрнутые группы можно развернуть кнопкой [+] над столбцами.").font = body; r += 1
    groups = [
        ("ID (A-I)", "Всегда видна", "Продукт, подпродукт, категория, инструмент, активность"),
        ("План (J-N)", "Видна", "Плановые даты 5 этапов (голубой фон)"),
        ("Факт (O-S)", "Видна", "Фактические даты 5 этапов (зелёный фон)"),
        ("Статус (T-V)", "Видна", "Текущий этап, прогресс %, RAG — автоформулы"),
        ("Анализ (W-AA)", "Свёрнута", "След. план, сдвиг дней, веса — автоформулы"),
        ("Базелайн (AB-AF)", "Скрыта", "Исходные плановые даты для сравнения"),
        ("Заметки (AG-AI)", "Свёрнута", "Эпики, комментарии, дата обновления"),
    ]
    for ci, h in enumerate(["Группа", "Видимость", "Содержание"]):
        c = ws.cell(r, ci+1, h); c.font = hs["font"]; c.fill = hs["fill"]; c.alignment = hs["alignment"]
    r += 1
    for name, vis, desc in groups:
        ws.cell(r, 1, name).font = Font(bold=True, size=11)
        ws.cell(r, 2, vis).font = body
        ws.cell(r, 3, desc).font = body
        r += 1
    r += 1

    # ── Section 5: FAQ ──
    ws.cell(r, 1, "5. FAQ").font = h2_font; r += 1
    faq = [
        ("Как добавить новый продукт?",
         "Добавьте строку в таблицу tblProducts на листе Справочники, затем перегенерируйте файл скриптом."),
        ("Пароль для редактирования формул?",
         '"edit" (без кавычек). Формульные столбцы защищены от случайного изменения.'),
        ("Почему строки неактивны?",
         'Столбец Активен = "Нет". Поставьте "Да" для нужных комбинаций.'),
        ("Как читать RAG?",
         "GREEN = в плане. AMBER = сдвиг 1-14 дней. RED = сдвиг >14 дней. DONE = 100% завершено."),
        ("Как фильтровать DATA?",
         "Используйте AutoFilter (стрелки в заголовках). Фильтруйте по продукту, категории или инструменту."),
    ]
    for q, a in faq:
        ws.cell(r, 1, q).font = h3_font; r += 1
        ws.cell(r, 1, a).font = body; r += 1
        r += 1


# ── Main ──────────────────────────────────────────────────────────────

def main():
    # Import exclusion/expansion rules from the dashboard generator
    from generate_dashboard import EXCLUDED_INSTRUMENTS, NEREZIDENTY_SUBPRODUCTS

    print("Parsing Confluence exports...")
    products = parse_products(ROADMAP_DIR / "Продукты+для+переезда.doc")
    instruments = parse_instruments(ROADMAP_DIR / "Инструменты+для+переезда.doc")

    # Apply same filters as generate_dashboard
    instruments = [i for i in instruments if i["instrument"] not in EXCLUDED_INSTRUMENTS]
    products = [p for p in products
                if not (p["prod"] == "Нерезиденты" and p["subprod"] == "Нерезиденты")
                ] + NEREZIDENTY_SUBPRODUCTS

    print(f"  Products: {len(products)} leaf nodes, Instruments: {len(instruments)}")

    wb = openpyxl.Workbook()
    print("Creating Справочники..."); create_справочники(wb, products, instruments)
    print("Creating DATA..."); _, total = create_data(wb, products, instruments)
    print("Creating Dashboard..."); create_dashboard(wb, products, instruments, total)
    print("Creating Category Dashboards..."); create_category_dashboards(wb, products, instruments)
    print("Creating Timeline..."); create_timeline(wb, products, instruments)
    print("Creating Gantt..."); create_gantt(wb, products, instruments)
    print("Creating Instructions..."); create_instructions(wb)

    wb.active = wb.sheetnames.index("Dashboard")
    print(f"Saving to {OUTPUT}..."); wb.save(str(OUTPUT))
    n_sheets = len(wb.sheetnames)
    print(f"Done! {total} data rows, {n_sheets} sheets.")

if __name__ == "__main__":
    main()
