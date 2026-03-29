#!/usr/bin/env python3
"""Generate interactive HTML dashboard for Переезд roadmap.

Self-contained script — no external imports from vault scripts.

Usage:
  python3 generate.py          # generate XLSX (if missing) + HTML
  python3 generate.py --init   # force-regenerate XLSX from Confluence
"""

import csv, json, datetime, re, sys, quopri
from html.parser import HTMLParser
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Paths ─────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent
SOURCE_DIR = ROOT / "source"
ENTRY_PATH = ROOT / "data_entry.xlsx"
HTML_PATH = ROOT / "docs" / "index.html"

# ── Constants ─────────────────────────────────────────────────────────
CATEGORIES = ["PRE", "1", "2", "3", "4"]
SUBSEGMENTS = [("Обычная (все грейсы)", [("до 200к", 0.5), ("свыше 200к", 0.5)])]

# 6-stage pipeline
STAGES = [("Старт разработки", 0.05), ("Интеграционное тестирование", 0.10),
          ("1%", 0.20), ("5%", 0.40), ("50%", 0.75), ("100%", 1.00)]

STAGE_NAMES = [s[0] for s in STAGES]
STAGE_WEIGHTS = dict(STAGES)
PLAN_COLS = [f"{s} план" for s in STAGE_NAMES]
FACT_COLS = [f"{s} факт" for s in STAGE_NAMES]
BASE_COLS = [f"{s} baseline" for s in STAGE_NAMES]
CSV_HEADERS = (["Агрегация","Продукт","Подпродукт","Подсегмент","Категория ПЗ",
                "Группа инструмента","Инструмент","Активен"]
               + PLAN_COLS + FACT_COLS + BASE_COLS + ["Эпики","Комментарии"])


# ── Confluence HTML Parser ───────────────────────────────────────────

class ConfluenceTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.in_table = self.in_row = self.in_cell = False
        self.rows, self._row, self._cell = [], [], ""
        self._col, self._rowspans, self._colspan, self._rowspan = 0, {}, 1, 1

    def handle_starttag(self, tag, attrs):
        a = dict(attrs)
        if tag == "table": self.in_table = True
        elif tag == "tr" and self.in_table:
            self.in_row = True; self._row = []; self._col = 0
        elif tag in ("td", "th") and self.in_row:
            self.in_cell = True; self._cell = ""
            while self._col in self._rowspans and self._rowspans[self._col] > 0:
                self._row.append(self._rowspans.get(f"{self._col}_v", ""))
                self._rowspans[self._col] -= 1
                if self._rowspans[self._col] == 0:
                    del self._rowspans[self._col]; self._rowspans.pop(f"{self._col}_v", None)
                self._col += 1
            self._colspan = int(a.get("colspan", "1")); self._rowspan = int(a.get("rowspan", "1"))

    def handle_data(self, data):
        if self.in_cell: self._cell += data.strip()

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self.in_cell:
            self.in_cell = False; v = self._cell.strip()
            for i in range(self._colspan):
                self._row.append(v if i == 0 else "")
                if self._rowspan > 1:
                    self._rowspans[self._col] = self._rowspan - 1
                    self._rowspans[f"{self._col}_v"] = v
                self._col += 1
        elif tag == "tr" and self.in_row:
            self.in_row = False
            while self._col in self._rowspans and self._rowspans[self._col] > 0:
                self._row.append(self._rowspans.get(f"{self._col}_v", ""))
                self._rowspans[self._col] -= 1
                if self._rowspans[self._col] == 0:
                    del self._rowspans[self._col]; self._rowspans.pop(f"{self._col}_v", None)
                self._col += 1
            self.rows.append(self._row)
        elif tag == "table": self.in_table = False


def _decode_confluence(path):
    raw = path.read_text(encoding="utf-8")
    s, e = raw.find("<html"), raw.find("</html>") + 7
    return quopri.decodestring(raw[s:e].encode()).decode("utf-8")

def _pct(s):
    s = s.strip().replace("%", "").replace(",", ".")
    try: return float(s) / 100.0 if s else 0.0
    except ValueError: return 0.0

def parse_products(path):
    html = _decode_confluence(path)
    p = ConfluenceTableParser(); p.feed(html)
    items = []
    for row in p.rows[1:]:
        while len(row) < 6: row.append("")
        sub = row[4].strip() if row[4].strip() else row[2].strip()
        items.append(dict(agg=row[0].strip(), w_agg=_pct(row[1]),
                          prod=row[2].strip(), w_prod=_pct(row[3]),
                          subprod=sub, w_subprod=_pct(row[5])))
    return items

def _clean_name(s):
    """Strip parenthetical notes from instrument names."""
    return re.sub(r'\s*\(.*$', '', s).strip()

def parse_instruments(path):
    html = _decode_confluence(path)
    body = html[html.find("<body"):html.find("</body>")]
    items = []
    for grp_html, sub_html in re.findall(r'<li><p[^>]*>(.*?)</p>\s*(?:<ol[^>]*>(.*?)</ol>)?', body, re.DOTALL):
        grp = _clean_name(re.sub(r'<[^>]+>', '', grp_html).strip())
        if sub_html.strip():
            subs = re.findall(r'<li>.*?<p[^>]*>(.*?)</p>.*?</li>', sub_html, re.DOTALL)
            if not subs: subs = re.findall(r'<li><span>(.*?)</span></li>', sub_html, re.DOTALL)
            for si in subs:
                items.append(dict(group=grp, instrument=_clean_name(re.sub(r'<[^>]+>', '', si).strip())))
        else:
            items.append(dict(group=grp, instrument=grp))
    return items


# ── XLSX Entry ───────────────────────────────────────────────────────

BLUE_H = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GREEN_H = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
YELLOW_H = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HDR_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HDR_FONT = Font(bold=True, size=10, color="FFFFFF")

def generate_entry_xlsx(products, instruments):
    """Generate a simple XLSX for data entry — no formulas, instant open."""
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "DATA"

    # Headers
    for ci, h in enumerate(CSV_HEADERS, 1):
        c = ws.cell(1, ci, h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Column widths and fills
    n_stages = len(STAGES)
    widths = [16,16,22,14,12,20,22,9] + [11]*n_stages + [11]*n_stages + [11]*n_stages + [30,25]
    plan_start = 8
    date_plan_cols = list(range(plan_start, plan_start + n_stages))
    date_fact_cols = list(range(plan_start + n_stages, plan_start + 2*n_stages))
    date_base_cols = list(range(plan_start + 2*n_stages, plan_start + 3*n_stages))
    for ci, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(ci+1)].width = w

    # Data
    sm = {sp: segs for sp, segs in SUBSEGMENTS}
    row_n = 2
    for p in products:
        for cat in CATEGORIES:
            for inst in instruments:
                base = [p["agg"], p["prod"], p["subprod"], "", cat,
                        inst["group"], inst["instrument"], "Нет"]
                entries = []
                if p["subprod"] in sm:
                    for sn, _ in sm[p["subprod"]]:
                        entries.append(base[:3] + [sn] + base[4:])
                else:
                    entries.append(base)
                for vals in entries:
                    for ci, v in enumerate(vals):
                        ws.cell(row_n, ci+1, v)
                    # Apply fills to date columns
                    for ci in date_plan_cols:
                        ws.cell(row_n, ci+1).fill = BLUE_H
                        ws.cell(row_n, ci+1).number_format = 'DD.MM.YY'
                    for ci in date_fact_cols:
                        ws.cell(row_n, ci+1).fill = GREEN_H
                        ws.cell(row_n, ci+1).number_format = 'DD.MM.YY'
                    for ci in date_base_cols:
                        ws.cell(row_n, ci+1).fill = YELLOW_H
                        ws.cell(row_n, ci+1).number_format = 'DD.MM.YY'
                    row_n += 1

    total = row_n - 2

    # Data validation: Активен = Да/Нет
    dv = DataValidation(type="list", formula1='"Да,Нет"', allow_blank=False)
    dv.add(f"H2:H{row_n-1}"); ws.add_data_validation(dv)

    # Freeze + AutoFilter
    ws.freeze_panes = "I2"
    ws.auto_filter.ref = f"A1:Y{row_n-1}"

    wb.save(str(ENTRY_PATH))
    print(f"  XLSX: {total} rows -> {ENTRY_PATH.name}")

def read_entry_xlsx():
    """Read the data entry XLSX."""
    wb = openpyxl.load_workbook(str(ENTRY_PATH), data_only=True)
    ws = wb.active
    headers = [ws.cell(1, ci).value for ci in range(1, ws.max_column + 1)]
    rows = []
    for ri in range(2, ws.max_row + 1):
        row = {}
        for ci, h in enumerate(headers, 1):
            v = ws.cell(ri, ci).value
            if isinstance(v, datetime.datetime):
                v = v.strftime("%Y-%m-%d")
            elif isinstance(v, datetime.date):
                v = v.isoformat()
            row[h] = str(v) if v is not None else ""
        rows.append(row)
    return rows


# ── Computation ──────────────────────────────────────────────────────

def _parse_date(s):
    if not s or not s.strip(): return None
    s = s.strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%y", "%d.%m.%Y", "%d/%m/%Y"):
        try: return datetime.datetime.strptime(s, fmt).date()
        except ValueError: continue
    return None

def _d2s(d):
    return d.strftime("%d.%m.%y") if d else ""

def _d2iso(d):
    return d.isoformat() if d else None

def compute(rows, products):
    prod_w = {}
    for p in products:
        k = f'{p["agg"]}|{p["prod"]}|{p["subprod"]}'
        prod_w[k] = p["w_agg"] * p["w_prod"] * p["w_subprod"]
    subseg_w = {f"{sp}|{sn}": sw for sp, segs in SUBSEGMENTS for sn, sw in segs}

    out = []
    for row in rows:
        r = {}
        r["agg"] = row["Агрегация"]
        r["prod"] = row["Продукт"]
        r["subprod"] = row["Подпродукт"]
        r["subseg"] = row.get("Подсегмент", "")
        r["cat"] = row["Категория ПЗ"]
        r["igrp"] = row["Группа инструмента"]
        r["instr"] = row["Инструмент"]
        r["seg"] = f'{r["prod"]} | {r["cat"]}'
        r["active"] = row.get("Активен", "").strip().lower() in ("да", "yes", "1")
        r["epics"] = row.get("Эпики", "")
        r["comment"] = row.get("Комментарии", "")

        # Weight
        k = f'{r["agg"]}|{r["prod"]}|{r["subprod"]}'
        w = prod_w.get(k, 0)
        if r["subseg"]:
            w *= subseg_w.get(f'{r["subprod"]}|{r["subseg"]}', 1)
        r["weight"] = w

        # Dates
        plans = [_parse_date(row.get(c, "")) for c in PLAN_COLS]
        facts = [_parse_date(row.get(c, "")) for c in FACT_COLS]
        bases = [_parse_date(row.get(c, "")) for c in BASE_COLS]
        r["plans"] = [_d2iso(d) for d in plans]
        r["facts"] = [_d2iso(d) for d in facts]

        # Stage & progress
        stage = "Не начат"
        for i in range(len(STAGES) - 1, -1, -1):
            if facts[i]:
                stage = STAGE_NAMES[i]; break
        r["stage"] = stage
        r["progress"] = STAGE_WEIGHTS.get(stage, 0)

        # Slippage & RAG — compare next plan date vs today
        today = datetime.date.today()
        n_stages = len(STAGES)
        next_plan = None
        for i in range(n_stages):
            if not facts[i] and plans[i]:
                next_plan = plans[i]; break
        slip = (today - next_plan).days if next_plan else None
        r["slip"] = slip
        if facts[n_stages - 1]:
            r["rag"] = "DONE"
        elif slip is None:
            r["rag"] = "—"
        elif slip > 14:
            r["rag"] = "RED"
        elif slip > 0:
            r["rag"] = "AMBER"
        else:
            r["rag"] = "GREEN"

        # Gantt dates (for chart)
        valid_plans = [d for d in plans if d]
        valid_facts = [d for d in facts if d]
        r["gantt_start"] = _d2iso(min(valid_plans)) if valid_plans else None
        r["gantt_end"] = _d2iso(max(valid_plans)) if valid_plans else None
        r["gantt_fact"] = _d2iso(max(valid_facts)) if valid_facts else None

        out.append(r)
    return out


# ── HTML ─────────────────────────────────────────────────────────────

def generate_html(data):
    active = [r for r in data if r["active"]]
    # KPIs
    total_w = sum(r["weight"] for r in active) or 1
    progress = sum(r["weight"] * r["progress"] for r in active) / total_w
    n_red = sum(1 for r in active if r["rag"] == "RED")
    n_done = sum(1 for r in active if r["rag"] == "DONE")
    n_active = len(active)

    # Unique values for filters
    products = sorted(set(r["prod"] for r in data))
    categories = CATEGORIES
    instruments = sorted(set(r["instr"] for r in data))
    segments = sorted(set(r["seg"] for r in data))

    json_data = json.dumps({
        "rows": data,
        "kpis": {"progress": round(progress, 4), "red": n_red, "done": n_done, "active": n_active},
        "products": products,
        "categories": categories,
        "instruments": instruments,
        "segments": segments,
        "stages": STAGE_NAMES,
    }, ensure_ascii=False, default=str)

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Roadmap Переезд</title>
<script src="https://cdn.plot.ly/plotly-2.35.0.min.js"></script>
<style>
:root {{
  --blue: #2F5496; --blue-l: #D6E4F0; --green: #70AD47; --green-l: #C6EFCE;
  --amber: #FFC000; --amber-l: #FFEB9C; --red: #C00000; --red-l: #FFC7CE;
  --gray: #F2F2F2; --border: #E0E0E0; --text: #333; --muted: #888;
}}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
       color: var(--text); background: #FAFAFA; }}
.header {{ background: var(--blue); color: #fff; padding: 16px 24px; }}
.header h1 {{ font-size: 22px; font-weight: 600; }}
.header .sub {{ font-size: 13px; opacity: .7; margin-top: 4px; }}
.kpi-strip {{ display: flex; gap: 16px; padding: 16px 24px; background: #fff;
              border-bottom: 1px solid var(--border); flex-wrap: wrap; }}
.kpi {{ text-align: center; padding: 12px 20px; border-radius: 8px; background: var(--gray);
        min-width: 120px; }}
.kpi .val {{ font-size: 28px; font-weight: 700; color: var(--blue); }}
.kpi .lbl {{ font-size: 11px; color: var(--muted); text-transform: uppercase; letter-spacing: .5px; }}
.kpi.red .val {{ color: var(--red); }}
.kpi.green .val {{ color: var(--green); }}
.controls {{ display: flex; gap: 12px; padding: 12px 24px; background: #fff;
             border-bottom: 1px solid var(--border); align-items: center; flex-wrap: wrap; }}
.controls label {{ font-size: 12px; color: var(--muted); }}
.dd {{ position: relative; display: inline-block; }}
.dd-btn {{ padding: 6px 28px 6px 10px; border: 1px solid var(--border); border-radius: 4px;
           font-size: 13px; background: #fff; cursor: pointer; min-width: 150px; text-align: left;
           white-space: nowrap; overflow: hidden; text-overflow: ellipsis; max-width: 220px;
           appearance: none; position: relative; }}
.dd-btn::after {{ content: '\u25be'; position: absolute; right: 8px; top: 50%; transform: translateY(-50%);
                  color: var(--muted); pointer-events: none; }}
.dd-btn.has-selection {{ border-color: var(--blue); color: var(--blue); font-weight: 600; }}
.dd-list {{ display: none; position: absolute; top: 100%; left: 0; z-index: 10; background: #fff;
            border: 1px solid var(--border); border-radius: 6px; box-shadow: 0 4px 12px rgba(0,0,0,.12);
            max-height: 280px; overflow-y: auto; min-width: 100%; margin-top: 2px; }}
.dd-list.open {{ display: block; }}
.dd-list label {{ display: flex; align-items: center; gap: 6px; padding: 6px 12px; cursor: pointer;
                  font-size: 13px; white-space: nowrap; }}
.dd-list label:hover {{ background: var(--blue-l); }}
.dd-list input[type=checkbox] {{ accent-color: var(--blue); }}
.dd-all {{ border-bottom: 1px solid var(--border); font-weight: 600; }}
.tabs {{ display: flex; gap: 0; background: #fff; border-bottom: 2px solid var(--border);
         padding: 0 24px; }}
.tab-btn {{ padding: 10px 20px; border: none; background: none; cursor: pointer;
            font-size: 14px; color: var(--muted); border-bottom: 2px solid transparent;
            margin-bottom: -2px; transition: .2s; }}
.tab-btn.active {{ color: var(--blue); border-bottom-color: var(--blue); font-weight: 600; }}
.tab-btn:hover {{ color: var(--blue); }}
.tab-content {{ display: none; padding: 20px 24px; }}
.tab-content.active {{ display: block; }}
table {{ border-collapse: collapse; width: 100%; font-size: 12px; }}
th {{ background: var(--blue); color: #fff; padding: 8px 6px; text-align: center;
     font-weight: 600; position: sticky; top: 0; z-index: 1; white-space: nowrap; }}
td {{ padding: 6px; border: 1px solid var(--border); text-align: center; white-space: nowrap; }}
tr:nth-child(even) {{ background: #FAFAFA; }}
tr:hover {{ background: var(--blue-l); }}
.rag-RED {{ background: var(--red-l); color: var(--red); font-weight: 700; }}
.rag-AMBER {{ background: var(--amber-l); color: #8B6914; font-weight: 700; }}
.rag-GREEN {{ background: var(--green-l); color: #2E7D32; font-weight: 700; }}
.rag-DONE {{ background: #BDD7EE; color: var(--blue); font-weight: 700; }}
.prog-bar {{ width: 60px; height: 16px; background: #E8E8E8; border-radius: 3px;
             display: inline-block; vertical-align: middle; overflow: hidden; }}
.prog-fill {{ height: 100%; background: var(--blue); border-radius: 3px; transition: .3s; }}
.matrix-wrap {{ overflow-x: auto; }}
.matrix td {{ min-width: 50px; font-size: 11px; }}
.matrix th.instr {{ writing-mode: vertical-lr; text-orientation: mixed; padding: 8px 4px;
                    font-size: 10px; min-width: 35px; }}
.stage-grp {{ background: #1B3A6B; font-size: 13px; }}
.plan-cell {{ background: var(--blue-l); }}
.fact-cell {{ background: var(--green-l); }}
.fact-done {{ background: var(--green); color: #fff; font-weight: 600; }}
.section-title {{ font-size: 16px; font-weight: 600; color: var(--blue); margin: 16px 0 8px; }}
.info-box {{ background: #fff; border: 1px solid var(--border); border-radius: 8px;
             padding: 20px; margin: 12px 0; line-height: 1.6; }}
.info-box h3 {{ color: var(--blue); margin-bottom: 8px; }}
.info-box ol {{ padding-left: 20px; }}
.info-box li {{ margin: 6px 0; }}
.legend {{ display: flex; gap: 16px; margin: 12px 0; font-size: 12px; flex-wrap: wrap; }}
.legend span {{ display: inline-flex; align-items: center; gap: 4px; }}
.legend i {{ width: 14px; height: 14px; border-radius: 3px; display: inline-block; }}
td.left {{ text-align: left; }}
.gantt-section {{ margin: 20px 0; }}
@media (max-width: 768px) {{
  .kpi-strip {{ gap: 8px; padding: 8px; }}
  .kpi {{ min-width: 80px; padding: 8px; }}
  .kpi .val {{ font-size: 20px; }}
  .controls {{ padding: 8px; }}
}}
.auth-overlay {{ position:fixed; inset:0; z-index:999; background:var(--blue);
  display:flex; align-items:center; justify-content:center; }}
.auth-box {{ background:#fff; border-radius:12px; padding:40px; text-align:center;
  box-shadow:0 8px 32px rgba(0,0,0,.2); min-width:300px; }}
.auth-box h2 {{ color:var(--blue); margin-bottom:16px; }}
.auth-box input {{ padding:10px 14px; border:1px solid var(--border); border-radius:6px;
  font-size:15px; width:100%; margin-bottom:12px; }}
.auth-box button {{ padding:10px 24px; background:var(--blue); color:#fff; border:none;
  border-radius:6px; font-size:14px; cursor:pointer; width:100%; }}
.auth-box button:hover {{ opacity:.9; }}
.auth-box .err {{ color:var(--red); font-size:12px; margin-top:8px; display:none; }}
</style>
</head>
<body>

<div class="auth-overlay" id="auth-overlay">
  <div class="auth-box">
    <h2>Roadmap Переезд</h2>
    <p style="color:var(--muted);font-size:13px;margin-bottom:16px">Введите пароль для доступа</p>
    <input type="password" id="auth-pw" placeholder="Пароль" onkeydown="if(event.key==='Enter')checkAuth()">
    <button onclick="checkAuth()">Войти</button>
    <div class="err" id="auth-err">Неверный пароль</div>
  </div>
</div>

<div id="app-content" style="display:none">
<div class="header">
  <h1>Roadmap «Переезд»</h1>
  <div class="sub">Миграция на целевую архитектуру — интерактивный дашборд</div>
</div>

<div class="kpi-strip" id="kpis"></div>

<div class="controls">
  <div><label>Продукт</label><br>
    <div class="dd" id="dd-prod"><button class="dd-btn" onclick="toggleDD('dd-prod')">Все</button><div class="dd-list" id="ddl-prod"></div></div></div>
  <div><label>Категория ПЗ</label><br>
    <div class="dd" id="dd-cat"><button class="dd-btn" onclick="toggleDD('dd-cat')">Все</button><div class="dd-list" id="ddl-cat"></div></div></div>
  <div><label>Инструмент</label><br>
    <div class="dd" id="dd-instr"><button class="dd-btn" onclick="toggleDD('dd-instr')">Все</button><div class="dd-list" id="ddl-instr"></div></div></div>
  <div style="display:flex;flex-direction:column;gap:8px;justify-content:center">
    <label style="font-size:13px;cursor:pointer"><input type="checkbox" id="f-hide-inactive" onchange="render()"> Скрыть неактивные</label>
    <button onclick="clearFilters()" style="padding:4px 12px;font-size:12px;cursor:pointer;border:1px solid var(--border);border-radius:4px;background:#fff">Сбросить фильтры</button></div>
  <div style="margin-left:auto;font-size:12px;color:var(--muted)">
    Обновлено: {datetime.date.today().strftime('%d.%m.%Y')}</div>
</div>

<div class="tabs">
  <button class="tab-btn active" data-tab="dashboard" onclick="switchTab('dashboard')">Dashboard</button>
  <button class="tab-btn" data-tab="timeline" onclick="switchTab('timeline')">Timeline</button>
  <button class="tab-btn" data-tab="gantt" onclick="switchTab('gantt')">Гант</button>
  <button class="tab-btn" data-tab="data" onclick="switchTab('data')">Данные</button>
  <button class="tab-btn" data-tab="help" onclick="switchTab('help')">Инструкция</button>
</div>

<div class="tab-content active" id="tab-dashboard"></div>
<div class="tab-content" id="tab-timeline"></div>
<div class="tab-content" id="tab-gantt"></div>
<div class="tab-content" id="tab-data"></div>
<div class="tab-content" id="tab-help">
  <div class="info-box">
    <h3>Рабочий процесс</h3>
    <ol>
      <li><b>Откройте data_entry.xlsx</b> в Excel — файл без формул, открывается мгновенно.</li>
      <li><b>Активируйте комбинации:</b> в столбце H (Активен) выберите «Да» из выпадающего списка.</li>
      <li><b>Введите плановые даты</b> этапов (голубые столбцы: ИТ план … 100% план).</li>
      <li><b>Обновляйте фактические даты</b> по мере прогресса (зелёные столбцы).</li>
      <li><b>Сохраните и перегенерируйте:</b> <code>python3 generate.py</code></li>
      <li><b>Откройте docs/index.html</b> в браузере — дашборд обновится.</li>
    </ol>
  </div>
  <div class="info-box">
    <h3>Вкладки</h3>
    <table>
      <tr><th style="text-align:left">Вкладка</th><th style="text-align:left">Назначение</th></tr>
      <tr><td class="left"><b>Dashboard</b></td><td class="left">Матрица Сегмент x Инструмент — общий прогресс</td></tr>
      <tr><td class="left"><b>Timeline</b></td><td class="left">Этапы миграции с план/факт датами</td></tr>
      <tr><td class="left"><b>Гант</b></td><td class="left">Диаграмма Ганта по продуктам (5 категорий ПЗ)</td></tr>
      <tr><td class="left"><b>Данные</b></td><td class="left">Полная таблица со всеми вычисленными полями</td></tr>
    </table>
  </div>
  <div class="info-box">
    <h3>RAG-статус</h3>
    <div class="legend">
      <span><i style="background:var(--green-l)"></i> GREEN — в плане</span>
      <span><i style="background:var(--amber-l)"></i> AMBER — сдвиг 1-14 дней</span>
      <span><i style="background:var(--red-l)"></i> RED — сдвиг &gt;14 дней</span>
      <span><i style="background:#BDD7EE"></i> DONE — 100% завершено</span>
    </div>
  </div>
  <div class="info-box">
    <h3>Перепланирование</h3>
    <p>Перед изменением плановых дат скопируйте текущие планы в столбцы baseline.
       Столбец «Сдвиг» покажет разницу с базелайном. Без базелайна RAG = «—».</p>
  </div>
</div>
</div><!-- /app-content -->

<script>
const PW_HASH = 'fa7496e4ae840306df41bd658800392e24db8ac4767159d6bdf8a23b28c44ea0';
async function sha256(s) {{
  const buf = await crypto.subtle.digest('SHA-256', new TextEncoder().encode(s));
  return [...new Uint8Array(buf)].map(b => b.toString(16).padStart(2,'0')).join('');
}}
async function checkAuth() {{
  const pw = document.getElementById('auth-pw').value;
  const hash = await sha256(pw);
  if (hash === PW_HASH) {{
    sessionStorage.setItem('auth','1');
    document.getElementById('auth-overlay').style.display='none';
    document.getElementById('app-content').style.display='block';
  }} else {{
    document.getElementById('auth-err').style.display='block';
  }}
}}
if (sessionStorage.getItem('auth')==='1') {{
  document.getElementById('auth-overlay').style.display='none';
  document.getElementById('app-content').style.display='block';
}}
</script>

<script>
const D = {json_data};
const STAGES = D.stages;
const EPOCH = new Date(2026, 0, 1);

// ── Dropdown Filters ──
function buildDD(listId, items, sortFn) {{
  const list = document.getElementById(listId);
  if (sortFn) items = [...items].sort(sortFn);
  items.forEach(v => {{
    const lbl = document.createElement('label');
    const cb = document.createElement('input');
    cb.type = 'checkbox'; cb.value = v; cb.checked = false;
    cb.onchange = () => {{ updateDDBtn(listId); render(); }};
    lbl.appendChild(cb); lbl.appendChild(document.createTextNode(v));
    list.appendChild(lbl);
  }});
}}

function toggleDD(ddId) {{
  const list = document.getElementById(ddId).querySelector('.dd-list');
  document.querySelectorAll('.dd-list.open').forEach(l => {{ if (l !== list) l.classList.remove('open'); }});
  list.classList.toggle('open');
}}

document.addEventListener('click', e => {{
  if (!e.target.closest('.dd')) document.querySelectorAll('.dd-list.open').forEach(l => l.classList.remove('open'));
}});

function getDDValues(listId) {{
  return [...document.getElementById(listId).querySelectorAll('input:checked')].map(cb => cb.value);
}}

function updateDDBtn(listId) {{
  const dd = document.getElementById(listId).closest('.dd');
  const btn = dd.querySelector('.dd-btn');
  const checked = getDDValues(listId);
  if (checked.length === 0) {{
    btn.textContent = 'Все'; btn.classList.remove('has-selection');
  }} else if (checked.length <= 2) {{
    btn.textContent = checked.join(', '); btn.classList.add('has-selection');
  }} else {{
    btn.textContent = checked.length + ' выбрано'; btn.classList.add('has-selection');
  }}
}}

function initFilters() {{
  buildDD('ddl-prod', D.products, (a,b) => prodOrd(a) - prodOrd(b));
  buildDD('ddl-cat', D.categories);
  buildDD('ddl-instr', D.instruments);
}}

function matchFilter(vals, v) {{
  return vals.length === 0 || vals.includes(v);
}}

function clearFilters() {{
  document.querySelectorAll('.dd-list input[type=checkbox]').forEach(cb => cb.checked = false);
  document.querySelectorAll('.dd-btn').forEach(b => {{ b.textContent = 'Все'; b.classList.remove('has-selection'); }});
  document.getElementById('f-hide-inactive').checked = false;
  render();
}}

function getFiltered() {{
  const fp = getDDValues('ddl-prod');
  const fc = getDDValues('ddl-cat');
  const fi = getDDValues('ddl-instr');
  return D.rows.filter(r => {{
    if (!r.active) return false;
    if (!matchFilter(fp, r.prod)) return false;
    if (!matchFilter(fc, r.cat)) return false;
    if (!matchFilter(fi, r.instr)) return false;
    return true;
  }});
}}

// ── Tabs ──
function switchTab(t) {{
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.toggle('active', b.dataset.tab === t));
  document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
  document.getElementById('tab-' + t).classList.add('active');
  render();
}}

// ── Product & category sort ──
const PROD_ORDER = {{
  'КК':0,'КН':1,'КЛ':2,'КНР':3,'КНО':4,'POS':5,'BNPL':6,'Долями+':7,'Кубышка':8,
  'Незалоги.Дабл':9,
  'Авто':11,'Недвижимость':12,'Залоги.Дабл':13,
  'Умершие':20,'Банкроты':21,'Нерезиденты':22,'3P':23,'Инсталлмент':24
}};
const CAT_ORDER = {{'PRE':0,'1':1,'2':2,'3':3,'4':4}};
function prodOrd(p) {{ return PROD_ORDER[p] ?? 50; }}
function catOrd(c) {{ return CAT_ORDER[c] ?? 9; }}
function cmpProdCat(pA,cA,pB,cB) {{
  const dp = prodOrd(pA) - prodOrd(pB);
  return dp !== 0 ? dp : catOrd(cA) - catOrd(cB);
}}

// ── Helpers ──
function fmtPct(v) {{ return (v * 100).toFixed(1) + '%'; }}
function fmtDate(s) {{
  if (!s) return '';
  const d = new Date(s);
  return String(d.getDate()).padStart(2,'0') + '.' + String(d.getMonth()+1).padStart(2,'0') + '.' + String(d.getFullYear()).slice(2);
}}
function ragClass(r) {{ return r && r !== '—' ? 'rag-' + r : ''; }}
function progBar(v) {{
  return `<div class="prog-bar"><div class="prog-fill" style="width:${{Math.round(v*100)}}%"></div></div> ${{fmtPct(v)}}`;
}}
function dateToDays(s) {{
  if (!s) return null;
  return Math.round((new Date(s) - EPOCH) / 86400000);
}}

// ── KPIs ──
function renderKPIs() {{
  const rows = getFiltered();
  const tw = rows.reduce((s, r) => s + r.weight, 0) || 1;
  const prog = rows.reduce((s, r) => s + r.weight * r.progress, 0) / tw;
  const red = rows.filter(r => r.rag === 'RED').length;
  const done = rows.filter(r => r.rag === 'DONE').length;
  document.getElementById('kpis').innerHTML = `
    <div class="kpi"><div class="val">${{fmtPct(prog)}}</div><div class="lbl">Прогресс</div></div>
    <div class="kpi"><div class="val">${{rows.length}}</div><div class="lbl">Активных</div></div>
    <div class="kpi red"><div class="val">${{red}}</div><div class="lbl">RED</div></div>
    <div class="kpi green"><div class="val">${{done}}</div><div class="lbl">DONE</div></div>`;
}}

// ── Dashboard matrix ──
function renderDashboard() {{
  const fp = getDDValues('ddl-prod');
  const fc = getDDValues('ddl-cat');
  const fi = getDDValues('ddl-instr');
  const allRows = D.rows.filter(r => {{
    if (!matchFilter(fp, r.prod)) return false;
    if (!matchFilter(fc, r.cat)) return false;
    if (!matchFilter(fi, r.instr)) return false;
    return true;
  }});
  const hideInactive = document.getElementById('f-hide-inactive').checked;
  let segs = [...new Set(allRows.map(r => r.seg))].sort((a,b) => {{
    const [pA,cA] = a.split(' | '); const [pB,cB] = b.split(' | ');
    return cmpProdCat(pA,cA,pB,cB);
  }});
  const instrs = [...new Set(allRows.map(r => r.instr))].sort();
  if (hideInactive) {{
    segs = segs.filter(seg => allRows.some(r => r.seg === seg && r.active));
  }}
  if (!segs.length) {{ document.getElementById('tab-dashboard').innerHTML = '<p>Нет данных</p>'; return; }}
  let h = '<div class="legend" style="margin-bottom:8px"><span><i style="background:rgba(47,84,150,0.15)"></i> Активно (0%)</span>'
    + '<span><i style="background:rgba(47,84,150,0.7)"></i> Активно (прогресс)</span>'
    + '<span><i style="background:#F0F0F0;border:1px solid #ddd"></i> Не предусмотрено</span></div>';
  const PROD_GRP = {{
    'КК':'Незалоговые','КН':'Незалоговые','КЛ':'Незалоговые','КНР':'Незалоговые','КНО':'Незалоговые',
    'POS':'Незалоговые','BNPL':'Незалоговые','Долями+':'Незалоговые','Кубышка':'Незалоговые','Незалоги.Дабл':'Незалоговые',
    'Авто':'Залоговые','Недвижимость':'Залоговые','Залоги.Дабл':'Залоговые',
    'Умершие':'Спецсегменты','Банкроты':'Спецсегменты','Нерезиденты':'Спецсегменты','3P':'Спецсегменты','Инсталлмент':'Спецсегменты'
  }};
  const nCols = instrs.length + 1;
  h += '<div class="matrix-wrap"><table class="matrix"><thead><tr><th>Сегмент</th>';
  instrs.forEach(i => h += `<th class="instr">${{i}}</th>`);
  h += '</tr></thead><tbody>';
  let lastGrp = '';
  segs.forEach(seg => {{
    const prod = seg.split(' | ')[0];
    const grp = PROD_GRP[prod] || 'Прочие';
    if (grp !== lastGrp) {{
      h += `<tr><td colspan="${{nCols}}" style="background:var(--blue);color:#fff;font-weight:700;font-size:13px;padding:6px 10px;text-align:left">${{grp}}</td></tr>`;
      lastGrp = grp;
    }}
    h += `<tr><td class="left" style="font-weight:600">${{seg}}</td>`;
    instrs.forEach(instr => {{
      const all = allRows.filter(r => r.seg === seg && r.instr === instr);
      const active = all.filter(r => r.active);
      if (!active.length) {{
        h += `<td style="background:#F0F0F0;color:#ccc;font-size:10px">—</td>`;
      }} else {{
        const tw = active.reduce((s, r) => s + r.weight, 0) || 1;
        const p = active.reduce((s, r) => s + r.weight * r.progress, 0) / tw;
        const alpha = 0.15 + p * 0.65;
        const bg = `rgba(47,84,150,${{alpha.toFixed(2)}})`;
        const fg = p > 0.4 ? '#fff' : '#333';
        h += `<td style="background:${{bg}};color:${{fg}}">${{p > 0 ? fmtPct(p) : '0%'}}</td>`;
      }}
    }});
    h += '</tr>';
  }});
  h += '</tbody></table></div>';
  document.getElementById('tab-dashboard').innerHTML = h;
}}

// ── Timeline ──
function renderTimeline() {{
  const rows = getFiltered();
  const hideInactive = document.getElementById('f-hide-inactive').checked;
  const detailed = document.getElementById('f-tl-detail') && document.getElementById('f-tl-detail').checked;

  let h = '<div style="margin-bottom:8px"><label style="font-size:13px;cursor:pointer">'
    + '<input type="checkbox" id="f-tl-detail" onchange="render()"> Детализация по подпродуктам</label></div>';

  if (detailed) {{
    // ── Detailed: one row per subproduct x instrument ──
    let detail = rows.map(r => ({{...r}}));
    if (hideInactive) detail = detail.filter(r => r.plans.some(Boolean) || r.facts.some(Boolean));
    detail.sort((a,b) => {{
      const pc = cmpProdCat(a.prod, a.cat, b.prod, b.cat);
      if (pc !== 0) return pc;
      if (a.subprod !== b.subprod) return a.subprod.localeCompare(b.subprod);
      return a.instr.localeCompare(b.instr);
    }});

    h += '<div style="overflow-x:auto"><table><thead><tr><th>Сегмент</th><th>Подпродукт</th><th>Подсегмент</th><th>Инструмент</th>';
    STAGES.forEach(s => h += `<th colspan="2" class="stage-grp">${{s}}</th>`);
    h += '<th>Прогресс</th><th>RAG</th></tr><tr><th></th><th></th><th></th><th></th>';
    STAGES.forEach(() => h += '<th style="font-size:10px">план</th><th style="font-size:10px">факт</th>');
    h += '<th></th><th></th></tr></thead><tbody>';

    detail.forEach(r => {{
      h += `<tr><td class="left" style="font-weight:600">${{r.seg}}</td>`;
      h += `<td class="left">${{r.subprod}}</td><td>${{r.subseg||''}}</td><td>${{r.instr}}</td>`;
      for (let i = 0; i < STAGES.length; i++) {{
        const pc = r.plans[i] ? 'plan-cell' : '';
        const fc = r.facts[i] ? 'fact-done' : '';
        h += `<td class="${{pc}}">${{fmtDate(r.plans[i])}}</td><td class="${{fc}}">${{fmtDate(r.facts[i])}}</td>`;
      }}
      h += `<td>${{progBar(r.progress)}}</td><td class="${{ragClass(r.rag)}}">${{r.rag}}</td></tr>`;
    }});
    h += '</tbody></table></div>';

  }} else {{
    // ── Aggregated: segment x instrument ──
    const groups = {{}};
    rows.forEach(r => {{
      const k = r.seg + '|' + r.instr;
      if (!groups[k]) groups[k] = {{ seg: r.seg, prod: r.prod, cat: r.cat, instr: r.instr, items: [] }};
      groups[k].items.push(r);
    }});
    let entries = Object.values(groups).sort((a,b) => {{
      const pc = cmpProdCat(a.prod, a.cat, b.prod, b.cat);
      return pc !== 0 ? pc : a.instr.localeCompare(b.instr);
    }});
    if (hideInactive) {{
      entries = entries.filter(g => g.items.some(r => r.plans.some(Boolean) || r.facts.some(Boolean)));
    }}

    h += '<table><thead><tr><th>Сегмент</th><th>Инструмент</th>';
    STAGES.forEach(s => h += `<th colspan="2" class="stage-grp">${{s}}</th>`);
    h += '<th>Прогресс</th><th>RAG</th></tr><tr><th></th><th></th>';
    STAGES.forEach(() => h += '<th style="font-size:10px">план</th><th style="font-size:10px">факт</th>');
    h += '<th></th><th></th></tr></thead><tbody>';

    entries.forEach(g => {{
      const tw = g.items.reduce((s,r) => s + r.weight, 0) || 1;
      const prog = g.items.reduce((s,r) => s + r.weight * r.progress, 0) / tw;
      const plans = STAGES.map((_,i) => {{
        const dates = g.items.map(r => r.plans[i]).filter(Boolean);
        return dates.length ? dates.sort()[0] : null;
      }});
      const facts = STAGES.map((_,i) => {{
        const dates = g.items.map(r => r.facts[i]).filter(Boolean);
        return dates.length ? dates.sort().pop() : null;
      }});
      const rags = g.items.map(r => r.rag);
      const rag = rags.includes('RED') ? 'RED' : rags.includes('AMBER') ? 'AMBER' :
                  rags.every(r => r === 'DONE') && rags.length ? 'DONE' : 'GREEN';

      h += `<tr><td class="left" style="font-weight:600">${{g.seg}}</td><td>${{g.instr}}</td>`;
      STAGES.forEach((_,i) => {{
        const pc = plans[i] ? 'plan-cell' : '';
        const fc = facts[i] ? 'fact-done' : '';
        h += `<td class="${{pc}}">${{fmtDate(plans[i])}}</td><td class="${{fc}}">${{fmtDate(facts[i])}}</td>`;
      }});
      h += `<td>${{progBar(prog)}}</td><td class="${{ragClass(rag)}}">${{rag}}</td></tr>`;
    }});
    h += '</tbody></table>';
  }}
  document.getElementById('tab-timeline').innerHTML = h;
}}

// ── Gantt ──
const STAGE_COLORS = ['#264653','#2A9D8F','#E9C46A','#F4A261','#E76F51','#E63946'];
const STAGE_COLORS_LIGHT = ['#264653AA','#2A9D8FAA','#E9C46AAA','#F4A261AA','#E76F51AA','#E63946AA'];

function renderGantt() {{
  const rows = getFiltered();
  const container = document.getElementById('tab-gantt');
  container.innerHTML = '';

  D.categories.forEach(cat => {{
    const catRows = rows.filter(r => r.cat === cat);
    if (!catRows.length) return;

    // Group by product, aggregate plan/fact per stage
    const prods = {{}};
    catRows.forEach(r => {{
      if (!prods[r.prod]) prods[r.prod] = {{ plans: STAGES.map(()=>null), facts: STAGES.map(()=>null) }};
      const p = prods[r.prod];
      for (let i = 0; i < STAGES.length; i++) {{
        const pd = dateToDays(r.plans[i]);
        const fd = dateToDays(r.facts[i]);
        if (pd !== null && (p.plans[i] === null || pd < p.plans[i])) p.plans[i] = pd;
        if (fd !== null && (p.facts[i] === null || fd > p.facts[i])) p.facts[i] = fd;
      }}
    }});

    // Build stage durations per product
    const labels = [];
    const gapArr = [];
    const stageDurs = STAGES.map(() => []);  // stageDurs[stageIdx][prodIdx]
    const stageDone = STAGES.map(() => []);  // true/false per product

    Object.entries(prods).sort((a,b) => prodOrd(a[0]) - prodOrd(b[0])).forEach(([prod, d]) => {{
      if (d.plans.every(v => v === null)) return;
      labels.push(prod);

      // Find first non-null plan as gap
      const firstPlan = d.plans.find(v => v !== null) || 0;
      gapArr.push(firstPlan);

      for (let i = 0; i < STAGES.length; i++) {{
        let dur = 0;
        if (d.plans[i] !== null) {{
          // Duration = next stage plan - this stage plan (or 30 days for last stage)
          let nextPlan = null;
          for (let j = i + 1; j < STAGES.length; j++) {{
            if (d.plans[j] !== null) {{ nextPlan = d.plans[j]; break; }}
          }}
          dur = nextPlan !== null ? nextPlan - d.plans[i] : 30;
          if (dur < 0) dur = 0;
        }}
        stageDurs[i].push(dur);
        stageDone[i].push(d.facts[i] !== null);
      }}
    }});

    if (!labels.length) return;

    const div = document.createElement('div');
    div.className = 'gantt-section';
    div.id = 'gantt-' + cat;
    container.appendChild(div);

    // Month ticks
    const tv = [], tt = [];
    const mNames = ['Янв','Фев','Мар','Апр','Май','Июн','Июл','Авг','Сен','Окт','Ноя','Дек'];
    for (let y = 2026; y <= 2028; y++)
      for (let m = 0; m < 12; m++) {{
        const days = Math.round((new Date(y, m, 1) - EPOCH) / 86400000);
        if (days >= -30 && days <= 800) {{ tv.push(days); tt.push(mNames[m] + "'" + String(y).slice(2)); }}
      }}

    // Traces: gap + one per stage
    const traces = [
      {{ type:'bar', orientation:'h', y:labels, x:gapArr, name:'', showlegend:false,
         marker:{{color:'rgba(0,0,0,0)'}}, hoverinfo:'skip' }}
    ];
    for (let i = 0; i < STAGES.length; i++) {{
      // Per-bar color: solid if done, lighter if not
      const colors = stageDone[i].map(done => done ? STAGE_COLORS[i % STAGE_COLORS.length] : STAGE_COLORS_LIGHT[i % STAGE_COLORS_LIGHT.length]);
      traces.push({{
        type:'bar', orientation:'h', y:labels, x:stageDurs[i],
        name: STAGES[i],
        marker:{{ color: colors }},
        hovertemplate: STAGES[i] + ': %{{x}} дн.<extra></extra>'
      }});
    }}

    const todayDays = Math.round((new Date() - EPOCH) / 86400000);
    Plotly.newPlot(div, traces, {{
      barmode:'stack', title:'ПЗ ' + cat,
      xaxis:{{ tickvals:tv, ticktext:tt, gridcolor:'#eee',
               range:[0, Math.round((new Date(2026,11,31)-EPOCH)/864e5)] }},
      yaxis:{{ autorange:'reversed' }},
      height: Math.max(labels.length * 32 + 120, 250),
      margin:{{ l:150, r:20, t:40, b:40 }},
      legend:{{ orientation:'h', y:-0.15 }},
      shapes: [{{
        type: 'line', x0: todayDays, x1: todayDays, y0: 0, y1: 1, yref: 'paper',
        line: {{ color: '#C00000', width: 2, dash: 'dash' }}
      }}],
      annotations: [{{
        x: todayDays, y: 1.02, yref: 'paper', text: 'Сегодня', showarrow: false,
        font: {{ size: 11, color: '#C00000' }}
      }}]
    }}, {{ responsive:true }});
  }});

  if (!container.children.length) container.innerHTML = '<p>Нет данных для Ганта (нет плановых дат)</p>';
}}

// ── Data table ──
function renderData() {{
  const hideEmpty = document.getElementById('f-hide-inactive').checked;
  let rows = getFiltered();
  if (hideEmpty) {{
    rows = rows.filter(r => r.plans.some(Boolean) || r.facts.some(Boolean));
  }}
  let h = `<div style="overflow-x:auto;max-width:100%"><table style="min-width:800px"><thead><tr>
    <th>Продукт</th><th>Подпродукт</th><th>Кат.</th><th>Инструмент</th>
    <th>Этап</th><th>Прогресс</th><th>RAG</th><th>Сдвиг</th>`;
  STAGES.forEach(s => h += `<th>${{s}} п</th><th>${{s}} ф</th>`);
  h += `<th>Эпики</th><th>Комментарии</th>
  </tr></thead><tbody>`;
  rows.forEach(r => {{
    h += `<tr>
      <td class="left">${{r.prod}}</td><td class="left">${{r.subprod}}</td>
      <td>${{r.cat}}</td><td>${{r.instr}}</td>
      <td>${{r.stage}}</td><td>${{progBar(r.progress)}}</td>
      <td class="${{ragClass(r.rag)}}">${{r.rag}}</td>
      <td>${{r.slip !== null ? r.slip : ''}}</td>`;
    for (let i = 0; i < STAGES.length; i++) {{
      h += `<td class="plan-cell">${{fmtDate(r.plans[i])}}</td>`;
      h += `<td class="${{r.facts[i] ? 'fact-done' : ''}}">${{fmtDate(r.facts[i])}}</td>`;
    }}
    h += `<td class="left">${{r.epics||''}}</td><td class="left">${{r.comment||''}}</td></tr>`;
  }});
  h += '</tbody></table></div>';
  document.getElementById('tab-data').innerHTML = h;
}}

// ── Render all ──
function render() {{
  renderKPIs();
  const active = document.querySelector('.tab-btn.active');
  const tab = active ? active.dataset.tab : 'dashboard';
  if (tab === 'dashboard') renderDashboard();
  else if (tab === 'timeline') renderTimeline();
  else if (tab === 'gantt') renderGantt();
  else if (tab === 'data') renderData();
}}

initFilters();
render();
</script>
</body>
</html>"""

    HTML_PATH.parent.mkdir(parents=True, exist_ok=True)
    HTML_PATH.write_text(html, encoding="utf-8")
    print(f"  HTML: {HTML_PATH.name} ({HTML_PATH.stat().st_size // 1024} KB)")


# ── Main ─────────────────────────────────────────────────────────────

def main():
    force_init = "--init" in sys.argv

    print("Parsing Confluence exports...")
    products = parse_products(SOURCE_DIR / "Продукты+для+переезда.doc")
    instruments = parse_instruments(SOURCE_DIR / "Инструменты+для+переезда.doc")
    print(f"  Products: {len(products)}, Instruments: {len(instruments)}")

    if force_init or not ENTRY_PATH.exists():
        print("Generating entry XLSX...")
        generate_entry_xlsx(products, instruments)
    else:
        print(f"  XLSX: reading {ENTRY_PATH.name}")

    rows = read_entry_xlsx()
    print(f"  Rows: {len(rows)}")

    print("Computing derived values...")
    data = compute(rows, products)

    print("Generating HTML dashboard...")
    generate_html(data)
    print("Done!")

if __name__ == "__main__":
    main()
# test trigger
