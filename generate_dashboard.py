#!/usr/bin/env python3
"""Generate interactive HTML dashboard for Переезд roadmap.

Usage:
  python3 generate_dashboard.py          # generate XLSX (if missing) + HTML
  python3 generate_dashboard.py --init   # force-regenerate XLSX from Confluence
"""

import csv, json, datetime, sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).parent))
from generate_roadmap import (parse_products, parse_instruments,
                               CATEGORIES, SUBSEGMENTS, ROADMAP_DIR)

# 6-stage pipeline
STAGES = [("Старт разработки", 0.05), ("Интеграционное тестирование", 0.10),
          ("1%", 0.20), ("5%", 0.40), ("50%", 0.75), ("100%", 1.00)]

ENTRY_PATH = ROADMAP_DIR / "data_entry.xlsx"
HTML_PATH = ROADMAP_DIR / "Roadmap_Переезд.html"

STAGE_NAMES = [s[0] for s in STAGES]
STAGE_WEIGHTS = dict(STAGES)
PLAN_COLS = [f"{s} план" for s in STAGE_NAMES]
FACT_COLS = [f"{s} факт" for s in STAGE_NAMES]
BASE_COLS = [f"{s} baseline" for s in STAGE_NAMES]
CSV_HEADERS = (["Агрегация","Продукт","Подпродукт","Подсегмент","Категория ПЗ",
                "Группа инструмента","Инструмент","Активен"]
               + PLAN_COLS + FACT_COLS + BASE_COLS + ["Эпики","Комментарии"])

# ── XLSX Entry ───────────────────────────────────────────────────────

BLUE_H = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GREEN_H = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
YELLOW_H = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HDR_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HDR_FONT = Font(bold=True, size=10, color="FFFFFF")

def generate_entry_xlsx(products, instruments):
    """Generate a simple XLSX for data entry — no formulas, instant open."""
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "DATA"

    # Headers
    for ci, h in enumerate(CSV_HEADERS, 1):
        c = ws.cell(1, ci, h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Column widths and fills
    n_stages = len(STAGES)
    widths = [16,16,22,14,12,20,22,9] + [11]*n_stages + [11]*n_stages + [11]*n_stages + [30,25]
    plan_start = 8
    date_plan_cols = list(range(plan_start, plan_start + n_stages))
    date_fact_cols = list(range(plan_start + n_stages, plan_start + 2*n_stages))
    date_base_cols = list(range(plan_start + 2*n_stages, plan_start + 3*n_stages))
    for ci, w in enumerate(widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci+1)].width = w

    # Data
    sm = {sp: segs for sp, segs in SUBSEGMENTS}
    row_n = 2
    for p in products:
        for cat in CATEGORIES:
            for inst in instruments:
                base = [p["agg"], p["prod"], p["subprod"], "", cat,
                        inst["group"], inst["instrument"], "Нет"]
                entries = []
                if p["subprod"] in sm:
                    for sn, _ in sm[p["subprod"]]:
                        entries.append(base[:3] + [sn] + base[4:])
                else:
                    entries.append(base)
                for vals in entries:
                    for ci, v in enumerate(vals):
                        ws.cell(row_n, ci+1, v)
                    # Apply fills to date columns
                    for ci in date_plan_cols:
                        ws.cell(row_n, ci+1).fill = BLUE_H
                        ws.cell(row_n, ci+1).number_format = 'DD.MM.YY'
                    for ci in date_fact_cols:
                        ws.cell(row_n, ci+1).fill = GREEN_H
                        ws.cell(row_n, ci+1).number_format = 'DD.MM.YY'
                    for ci in date_base_cols:
                        ws.cell(row_n, ci+1).fill = YELLOW_H
                        ws.cell(row_n, ci+1).number_format = 'DD.MM.YY'
                    row_n += 1

    total = row_n - 2

    # Data validation: Активен = Да/Нет
    dv = DataValidation(type="list", formula1='"Да,Нет"', allow_blank=False)
    dv.add(f"H2:H{row_n-1}"); ws.add_data_validation(dv)

    # Freeze + AutoFilter
    ws.freeze_panes = "I2"
    ws.auto_filter.ref = f"A1:Y{row_n-1}"

    wb.save(str(ENTRY_PATH))
    print(f"  XLSX: {total} rows → {ENTRY_PATH.name}")

def read_entry_xlsx():
    """Read the data entry XLSX."""
    wb = openpyxl.load_workbook(str(ENTRY_PATH), data_only=True)
    ws = wb.active
    headers = [ws.cell(1, ci).value for ci in range(1, ws.max_column + 1)]
    rows = []
    for ri in range(2, ws.max_row + 1):
        row = {}
        for ci, h in enumerate(headers, 1):
            v = ws.cell(ri, ci).value
            if isinstance(v, datetime.datetime):
                v = v.strftime("%Y-%m-%d")
            elif isinstance(v, datetime.date):
                v = v.isoformat()
            row[h] = str(v) if v is not None else ""
        rows.append(row)
    return rows

# ── Computation ──────────────────────────────────────────────────────

def _parse_date(s):
    if not s or not s.strip(): return None
    s = s.strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%y", "%d.%m.%Y", "%d/%m/%Y"):
        try: return datetime.datetime.strptime(s, fmt).date()
        except ValueError: continue
    return None

def _d2s(d):
    return d.strftime("%d.%m.%y") if d else ""

def _d2iso(d):
    return d.isoformat() if d else None

def compute(rows, products):
    prod_w = {}
    for p in products:
        k = f'{p["agg"]}|{p["prod"]}|{p["subprod"]}'
        prod_w[k] = p["w_agg"] * p["w_prod"] * p["w_subprod"]
    subseg_w = {f"{sp}|{sn}": sw for sp, segs in SUBSEGMENTS for sn, sw in segs}

    out = []
    for row in rows:
        r = {}
        r["agg"] = row["Агрегация"]
        r["prod"] = row["Продукт"]
        r["subprod"] = row["Подпродукт"]
        r["subseg"] = row.get("Подсегмент", "")
        r["cat"] = row["Категория ПЗ"]
        r["igrp"] = row["Группа инструмента"]
        r["instr"] = row["Инструмент"]
        r["seg"] = f'{r["prod"]} | {r["cat"]}'
        r["active"] = row.get("Активен", "").strip().lower() in ("да", "yes", "1")
        r["epics"] = row.get("Эпики", "")
        r["comment"] = row.get("Комментарии", "")

        # Weight
        k = f'{r["agg"]}|{r["prod"]}|{r["subprod"]}'
        w = prod_w.get(k, 0)
        if r["subseg"]:
            w *= subseg_w.get(f'{r["subprod"]}|{r["subseg"]}', 1)
        r["weight"] = w

        # Dates
        plans = [_parse_date(row.get(c, "")) for c in PLAN_COLS]
        facts = [_parse_date(row.get(c, "")) for c in FACT_COLS]
        bases = [_parse_date(row.get(c, "")) for c in BASE_COLS]
        r["plans"] = [_d2iso(d) for d in plans]
        r["facts"] = [_d2iso(d) for d in facts]

        # Stage & progress
        stage = "Не начат"
        for i in range(4, -1, -1):
            if facts[i]:
                stage = STAGE_NAMES[i]; break
        r["stage"] = stage
        r["progress"] = STAGE_WEIGHTS.get(stage, 0)

        # Slippage & RAG — compare next plan date vs today
        today = datetime.date.today()
        n_stages = len(STAGES)
        next_plan = None
        for i in range(n_stages):
            if not facts[i] and plans[i]:
                next_plan = plans[i]; break
        slip = (today - next_plan).days if next_plan else None
        r["slip"] = slip
        if facts[n_stages - 1]:
            r["rag"] = "DONE"
        elif slip is None:
            r["rag"] = "—"
        elif slip > 14:
            r["rag"] = "RED"
        elif slip > 0:
            r["rag"] = "AMBER"
        else:
            r["rag"] = "GREEN"

        # Gantt dates (for chart)
        valid_plans = [d for d in plans if d]
        valid_facts = [d for d in facts if d]
        r["gantt_start"] = _d2iso(min(valid_plans)) if valid_plans else None
        r["gantt_end"] = _d2iso(max(valid_plans)) if valid_plans else None
        r["gantt_fact"] = _d2iso(max(valid_facts)) if valid_facts else None

        out.append(r)
    return out

# ── HTML ─────────────────────────────────────────────────────────────

def generate_html(data):
    active = [r for r in data if r["active"]]
    # KPIs
    total_w = sum(r["weight"] for r in active) or 1
    progress = sum(r["weight"] * r["progress"] for r in active) / total_w
    n_red = sum(1 for r in active if r["rag"] == "RED")
    n_done = sum(1 for r in active if r["rag"] == "DONE")
    n_active = len(active)

    # Unique values for filters
    products = sorted(set(r["prod"] for r in data))
    categories = CATEGORIES
    instruments = sorted(set(r["instr"] for r in data))
    segments = sorted(set(r["seg"] for r in data))

    json_data = json.dumps({
        "rows": data,
        "kpis": {"progress": round(progress, 4), "red": n_red, "done": n_done, "active": n_active},
        "products": products,
        "categories": categories,
        "instruments": instruments,
        "segments": segments,
        "stages": STAGE_NAMES,
    }, ensure_ascii=False, default=str)

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Roadmap Переезд</title>
<script src="https://cdn.plot.ly/plotly-2.35.0.min.js"></script>
<script src="https://cdn.sheetjs.com/xlsx-0.20.3/package/dist/xlsx.full.min.js"></script>
<style>
:root {{
  --blue: #2F5496; --blue-l: #D6E4F0; --green: #70AD47; --green-l: #C6EFCE;
  --amber: #FFC000; --amber-l: #FFEB9C; --red: #C00000; --red-l: #FFC7CE;
  --gray: #F2F2F2; --border: #E0E0E0; --text: #333; --muted: #888;
}}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
       color: var(--text); background: #FAFAFA; }}
.header {{ background: var(--blue); color: #fff; padding: 16px 24px; }}
.header h1 {{ font-size: 22px; font-weight: 600; }}
.header .sub {{ font-size: 13px; opacity: .7; margin-top: 4px; }}
.kpi-strip {{ display: flex; gap: 16px; padding: 16px 24px; background: #fff;
              border-bottom: 1px solid var(--border); flex-wrap: wrap; }}
.kpi {{ text-align: center; padding: 12px 20px; border-radius: 8px; background: var(--gray);
        min-width: 120px; }}
.kpi .val {{ font-size: 28px; font-weight: 700; color: var(--blue); }}
.kpi .lbl {{ font-size: 11px; color: var(--muted); text-transform: uppercase; letter-spacing: .5px; }}
.kpi.red .val {{ color: var(--red); }}
.kpi.green .val {{ color: var(--green); }}
.controls {{ display: flex; gap: 12px; padding: 12px 24px; background: #fff;
             border-bottom: 1px solid var(--border); align-items: center; flex-wrap: wrap; }}
.controls label {{ font-size: 12px; color: var(--muted); }}
.dd {{ position: relative; display: inline-block; }}
.dd-btn {{ padding: 6px 28px 6px 10px; border: 1px solid var(--border); border-radius: 4px;
           font-size: 13px; background: #fff; cursor: pointer; min-width: 150px; text-align: left;
           white-space: nowrap; overflow: hidden; text-overflow: ellipsis; max-width: 220px;
           appearance: none; position: relative; }}
.dd-btn::after {{ content: '▾'; position: absolute; right: 8px; top: 50%; transform: translateY(-50%);
                  color: var(--muted); pointer-events: none; }}
.dd-btn.has-selection {{ border-color: var(--blue); color: var(--blue); font-weight: 600; }}
.dd-list {{ display: none; position: absolute; top: 100%; left: 0; z-index: 10; background: #fff;
            border: 1px solid var(--border); border-radius: 6px; box-shadow: 0 4px 12px rgba(0,0,0,.12);
            max-height: 280px; overflow-y: auto; min-width: 100%; margin-top: 2px; }}
.dd-list.open {{ display: block; }}
.dd-list label {{ display: flex; align-items: center; gap: 6px; padding: 6px 12px; cursor: pointer;
                  font-size: 13px; white-space: nowrap; }}
.dd-list label:hover {{ background: var(--blue-l); }}
.dd-list input[type=checkbox] {{ accent-color: var(--blue); }}
.dd-all {{ border-bottom: 1px solid var(--border); font-weight: 600; }}
.tabs {{ display: flex; gap: 0; background: #fff; border-bottom: 2px solid var(--border);
         padding: 0 24px; }}
.tab-btn {{ padding: 10px 20px; border: none; background: none; cursor: pointer;
            font-size: 14px; color: var(--muted); border-bottom: 2px solid transparent;
            margin-bottom: -2px; transition: .2s; }}
.tab-btn.active {{ color: var(--blue); border-bottom-color: var(--blue); font-weight: 600; }}
.tab-btn:hover {{ color: var(--blue); }}
.tab-content {{ display: none; padding: 20px 24px; }}
.tab-content.active {{ display: block; }}
table {{ border-collapse: collapse; width: 100%; font-size: 12px; }}
th {{ background: var(--blue); color: #fff; padding: 8px 6px; text-align: center;
     font-weight: 600; position: sticky; top: 0; z-index: 1; white-space: nowrap; }}
td {{ padding: 6px; border: 1px solid var(--border); text-align: center; white-space: nowrap; }}
tr:nth-child(even) {{ background: #FAFAFA; }}
tr:hover {{ background: var(--blue-l); }}
.rag-RED {{ background: var(--red-l); color: var(--red); font-weight: 700; }}
.rag-AMBER {{ background: var(--amber-l); color: #8B6914; font-weight: 700; }}
.rag-GREEN {{ background: var(--green-l); color: #2E7D32; font-weight: 700; }}
.rag-DONE {{ background: #BDD7EE; color: var(--blue); font-weight: 700; }}
.prog-bar {{ width: 60px; height: 16px; background: #E8E8E8; border-radius: 3px;
             display: inline-block; vertical-align: middle; overflow: hidden; }}
.prog-fill {{ height: 100%; background: var(--blue); border-radius: 3px; transition: .3s; }}
.matrix-wrap {{ overflow-x: auto; }}
.matrix td {{ min-width: 50px; font-size: 11px; }}
.matrix th.instr {{ writing-mode: vertical-lr; text-orientation: mixed; padding: 8px 4px;
                    font-size: 10px; min-width: 35px; }}
.stage-grp {{ background: #1B3A6B; font-size: 13px; }}
.plan-cell {{ background: var(--blue-l); }}
.fact-cell {{ background: var(--green-l); }}
.fact-done {{ background: var(--green); color: #fff; font-weight: 600; }}
.section-title {{ font-size: 16px; font-weight: 600; color: var(--blue); margin: 16px 0 8px; }}
.info-box {{ background: #fff; border: 1px solid var(--border); border-radius: 8px;
             padding: 20px; margin: 12px 0; line-height: 1.6; }}
.info-box h3 {{ color: var(--blue); margin-bottom: 8px; }}
.info-box ol {{ padding-left: 20px; }}
.info-box li {{ margin: 6px 0; }}
.legend {{ display: flex; gap: 16px; margin: 12px 0; font-size: 12px; flex-wrap: wrap; }}
.legend span {{ display: inline-flex; align-items: center; gap: 4px; }}
.legend i {{ width: 14px; height: 14px; border-radius: 3px; display: inline-block; }}
td.left {{ text-align: left; }}
.gantt-section {{ margin: 20px 0; }}
.ed-section {{ background: #fff; border: 1px solid var(--border); border-radius: 8px;
              padding: 16px; margin-bottom: 12px; }}
.ed-section h4 {{ color: var(--blue); margin-bottom: 10px; font-size: 14px; }}
.ed-filters {{ display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 8px; }}
.ed-filters .dd {{ min-width: 160px; }}
.ed-actions {{ display: flex; gap: 16px; align-items: flex-end; flex-wrap: wrap; margin: 12px 0; }}
.ed-actions input[type=date] {{ padding: 6px 10px; border: 1px solid var(--border); border-radius: 4px; font-size: 13px; }}
.ed-actions select {{ padding: 6px 10px; border: 1px solid var(--border); border-radius: 4px; font-size: 13px; }}
.ed-btn {{ padding: 8px 20px; border: none; border-radius: 6px; font-size: 13px; cursor: pointer; font-weight: 600; }}
.ed-btn-primary {{ background: var(--blue); color: #fff; }}
.ed-btn-primary:hover {{ opacity: .9; }}
.ed-btn-success {{ background: var(--green); color: #fff; }}
.ed-btn-success:hover {{ opacity: .9; }}
.ed-count {{ font-size: 14px; font-weight: 600; color: var(--blue); padding: 8px 0; }}
.ed-preview {{ max-height: 400px; overflow: auto; }}
.info-tip {{ display: inline-block; width: 14px; height: 14px; line-height: 14px; text-align: center;
             border-radius: 50%; background: var(--border); color: var(--muted); font-size: 10px;
             cursor: help; vertical-align: middle; margin-left: 2px; position: relative; }}
.tab-controls {{ display: flex; align-items: center; gap: 16px; padding: 8px 0 12px;
                 border-bottom: 1px solid var(--border); margin-bottom: 12px; flex-wrap: wrap; }}
.tab-controls:empty {{ display: none; }}
.tab-controls label {{ font-size: 13px; cursor: pointer; }}
.tab-controls select {{ padding: 5px 10px; border: 1px solid var(--border); border-radius: 4px; font-size: 13px; }}
@media (max-width: 768px) {{
  .kpi-strip {{ gap: 8px; padding: 8px; }}
  .kpi {{ min-width: 80px; padding: 8px; }}
  .kpi .val {{ font-size: 20px; }}
  .controls {{ padding: 8px; }}
}}
.auth-overlay {{ position:fixed; inset:0; z-index:999; background:var(--blue);
  display:flex; align-items:center; justify-content:center; }}
.auth-box {{ background:#fff; border-radius:12px; padding:40px; text-align:center;
  box-shadow:0 8px 32px rgba(0,0,0,.2); min-width:300px; }}
.auth-box h2 {{ color:var(--blue); margin-bottom:16px; }}
.auth-box input {{ padding:10px 14px; border:1px solid var(--border); border-radius:6px;
  font-size:15px; width:100%; margin-bottom:12px; }}
.auth-box button {{ padding:10px 24px; background:var(--blue); color:#fff; border:none;
  border-radius:6px; font-size:14px; cursor:pointer; width:100%; }}
.auth-box button:hover {{ opacity:.9; }}
.auth-box .err {{ color:var(--red); font-size:12px; margin-top:8px; display:none; }}
</style>
</head>
<body>

<div class="auth-overlay" id="auth-overlay">
  <div class="auth-box">
    <h2>Roadmap Переезд</h2>
    <p style="color:var(--muted);font-size:13px;margin-bottom:16px">Введите пароль для доступа</p>
    <input type="password" id="auth-pw" placeholder="Пароль" onkeydown="if(event.key==='Enter')checkAuth()">
    <button onclick="checkAuth()">Войти</button>
    <div class="err" id="auth-err">Неверный пароль</div>
  </div>
</div>

<div id="app-content" style="display:none">
<div class="header">
  <h1>Roadmap «Переезд»</h1>
  <div class="sub">Миграция на целевую архитектуру — интерактивный дашборд</div>
</div>

<div class="kpi-strip" id="kpis"></div>

<div class="controls">
  <div><label>Продукт</label><br>
    <div class="dd" id="dd-prod"><button class="dd-btn" onclick="toggleDD('dd-prod')">Все</button><div class="dd-list" id="ddl-prod"></div></div></div>
  <div><label>Категория ПЗ</label><br>
    <div class="dd" id="dd-cat"><button class="dd-btn" onclick="toggleDD('dd-cat')">Все</button><div class="dd-list" id="ddl-cat"></div></div></div>
  <div><label>Инструмент</label><br>
    <div class="dd" id="dd-instr"><button class="dd-btn" onclick="toggleDD('dd-instr')">Все</button><div class="dd-list" id="ddl-instr"></div></div></div>
  <div style="display:flex;flex-direction:column;gap:8px;justify-content:center">
    <label style="font-size:13px;cursor:pointer"><input type="checkbox" id="f-hide-inactive" onchange="render()" checked> Скрыть неактивные</label>
    <button onclick="clearFilters()" style="padding:4px 12px;font-size:12px;cursor:pointer;border:1px solid var(--border);border-radius:4px;background:#fff">Сбросить фильтры</button></div>
  <div style="margin-left:auto;font-size:12px;color:var(--muted)">
    Обновлено: {datetime.date.today().strftime('%d.%m.%Y')}</div>
</div>

<div class="tabs">
  <button class="tab-btn active" data-tab="dashboard" onclick="switchTab('dashboard')">Dashboard</button>
  <button class="tab-btn" data-tab="timeline" onclick="switchTab('timeline')">Timeline</button>
  <button class="tab-btn" data-tab="gantt" onclick="switchTab('gantt')">Гант</button>
  <button class="tab-btn" data-tab="data" onclick="switchTab('data')">Данные</button>
  <button class="tab-btn" data-tab="editor" onclick="switchTab('editor')">Редактор</button>
  <button class="tab-btn" data-tab="help" onclick="switchTab('help')">Инструкция</button>
</div>

<div class="tab-content active" id="tab-dashboard">
  <div class="tab-controls" id="dash-ctrl"></div>
  <div id="dash-body"></div>
</div>
<div class="tab-content" id="tab-timeline">
  <div class="tab-controls" id="tl-ctrl"></div>
  <div id="tl-body"></div>
</div>
<div class="tab-content" id="tab-gantt">
  <div class="tab-controls" id="gantt-ctrl"></div>
  <div id="gantt-body"></div>
</div>
<div class="tab-content" id="tab-data">
  <div class="tab-controls" id="data-ctrl"></div>
  <div id="data-body"></div>
</div>
<div class="tab-content" id="tab-editor">
  <div class="tab-controls" id="editor-ctrl"></div>
  <div id="editor-body"></div>
</div>

<div class="tab-content" id="tab-help">

  <div class="info-box">
    <h3>Описание вкладок</h3>
    <table>
      <tr><th style="text-align:left;width:120px">Вкладка</th><th style="text-align:left">Что показывает</th><th style="text-align:left">Когда использовать</th></tr>
      <tr><td class="left"><b>Dashboard</b></td>
        <td class="left">Матрица «Сегмент × Инструмент». Каждая ячейка — средневзвешенный прогресс по подпродуктам. Серые ячейки «—» = комбинация не предусмотрена (Активен = Нет).</td>
        <td class="left">Общая картина: где идёт работа, где пробелы. Статус-митинг.</td></tr>
      <tr><td class="left"><b>Timeline</b></td>
        <td class="left">Таблица с план/факт датами по каждому из 6 этапов. Агрегация по сегменту × инструменту, с переключателем на детализацию по подпродуктам.</td>
        <td class="left">Контроль конкретных сроков. Поиск просроченных этапов.</td></tr>
      <tr><td class="left"><b>Гант</b></td>
        <td class="left">Горизонтальная диаграмма Ганта. Каждый этап — свой цвет. Три режима: план, факт, сравнение. Детализация по продукту или подпродукту. Красная линия = сегодня.</td>
        <td class="left">Визуальный обзор таймлайна. Презентации, отчёты руководству.</td></tr>
      <tr><td class="left"><b>Данные</b></td>
        <td class="left">Полная таблица: все активные строки со всеми вычисленными полями (этап, прогресс, RAG, сдвиг, даты). Горизонтальная прокрутка.</td>
        <td class="left">Детальный анализ, поиск конкретных комбинаций, экспорт.</td></tr>
    </table>
  </div>

  <div class="info-box">
    <h3>RAG-статусы</h3>
    <p style="margin-bottom:12px">RAG определяется для каждой строки (подпродукт × категория ПЗ × инструмент) путём сравнения <b>ближайшего незакрытого планового этапа</b> с <b>сегодняшней датой</b>.</p>
    <table>
      <tr><th style="text-align:left;width:90px">Статус</th><th style="text-align:left">Условие</th><th style="text-align:left">Что означает</th></tr>
      <tr><td><span style="background:var(--green-l);padding:2px 8px;border-radius:3px;font-weight:700;color:#2E7D32">GREEN</span></td>
        <td class="left">Плановая дата ближайшего этапа ≥ сегодня (сдвиг ≤ 0 дней)</td>
        <td class="left">Всё идёт по плану, сроки не нарушены.</td></tr>
      <tr><td><span style="background:var(--amber-l);padding:2px 8px;border-radius:3px;font-weight:700;color:#8B6914">AMBER</span></td>
        <td class="left">Плановая дата просрочена на 1–14 дней</td>
        <td class="left">Небольшая задержка. Нужно внимание — возможно, требуется эскалация.</td></tr>
      <tr><td><span style="background:var(--red-l);padding:2px 8px;border-radius:3px;font-weight:700;color:var(--red)">RED</span></td>
        <td class="left">Плановая дата просрочена более чем на 14 дней</td>
        <td class="left">Критическая задержка. Требуется вмешательство и корректировка сроков.</td></tr>
      <tr><td><span style="background:#BDD7EE;padding:2px 8px;border-radius:3px;font-weight:700;color:var(--blue)">DONE</span></td>
        <td class="left">Фактическая дата последнего этапа (100%) заполнена</td>
        <td class="left">Миграция этой комбинации полностью завершена.</td></tr>
      <tr><td><span style="padding:2px 8px;border-radius:3px;color:var(--muted)">—</span></td>
        <td class="left">Нет плановых дат ни на одном этапе</td>
        <td class="left">Даты ещё не заведены. Строка не участвует в расчёте прогресса.</td></tr>
    </table>
  </div>

  <div class="info-box">
    <h3>Как считается прогресс</h3>
    <table>
      <tr><th style="text-align:left;width:200px">Этап</th><th style="text-align:left;width:80px">Вес</th><th style="text-align:left">Пояснение</th></tr>
      <tr><td class="left">Старт разработки</td><td>5%</td><td class="left">Разработка начата</td></tr>
      <tr><td class="left">Интеграционное тестирование</td><td>10%</td><td class="left">e2e-тесты запущены</td></tr>
      <tr><td class="left">1% раскатка</td><td>20%</td><td class="left">Пилот на 1% трафика</td></tr>
      <tr><td class="left">5% раскатка</td><td>40%</td><td class="left">Расширение пилота</td></tr>
      <tr><td class="left">50% раскатка</td><td>75%</td><td class="left">Половина трафика на новой архитектуре</td></tr>
      <tr><td class="left">100% раскатка</td><td>100%</td><td class="left">Полная миграция завершена</td></tr>
    </table>
    <p style="margin-top:12px"><b>Прогресс строки</b> = вес последнего завершённого этапа (у которого заполнена фактическая дата). Если ни один этап не завершён — 0%.</p>
    <p><b>Прогресс сегмента</b> (на Dashboard/Timeline) = средневзвешенное по весам подпродуктов. Веса берутся из справочника продуктов (source/Продукты+для+переезда.doc).</p>
    <p><b>Общий прогресс</b> (KPI-карточка) = средневзвешенное по всем активным строкам с учётом фильтров.</p>
  </div>

  <div class="info-box">
    <h3>Особенности подсчётов</h3>
    <ul style="padding-left:20px">
      <li><b>Неактивные строки</b> (Активен = Нет) не участвуют в расчёте прогресса и RAG. Они отображаются серым «—» на Dashboard.</li>
      <li><b>Строки без плановых дат</b> — RAG = «—», не влияют на KPI RED/AMBER/GREEN.</li>
      <li><b>Агрегация на Timeline</b> — план = самая ранняя дата среди подпродуктов (MIN), факт = самая поздняя (MAX). Этап считается закрытым на уровне сегмента только когда ВСЕ подпродукты его завершили.</li>
      <li><b>RAG на Timeline</b> — берётся наихудший статус среди подпродуктов: RED &gt; AMBER &gt; GREEN. DONE только если все подпродукты = DONE.</li>
      <li><b>Гант (факт)</b> — если этап завершён, но следующий ещё не начат, бар текущего этапа тянется до сегодняшней даты (идёт работа).</li>
      <li><b>Подсегменты</b> — КК Обычная делится на «до 200к» и «свыше 200к» (вес 50/50). Остальные продукты без подсегментов.</li>
      <li><b>Фильтры</b> — применяются ко всем вкладкам одновременно. Если ничего не выбрано — показываются все. Выбор нескольких = логическое ИЛИ внутри фильтра.</li>
    </ul>
  </div>

  <div class="info-box">
    <h3>Как обновлять данные</h3>
    <ol>
      <li><b>Откройте data_entry.xlsx</b> в Excel — файл без формул, открывается мгновенно.</li>
      <li><b>Активируйте комбинации:</b> в столбце H (Активен) выберите «Да» из выпадающего списка.</li>
      <li><b>Введите плановые даты</b> (голубые столбцы) и фактические (зелёные столбцы).</li>
      <li><b>Сохраните</b> файл и запустите <code>python3 generate.py</code></li>
      <li><b>Закоммитьте:</b> <code>git add -A &amp;&amp; git commit -m "Обновление" &amp;&amp; git push</code></li>
      <li>Через ~1 минуту дашборд обновится на GitHub Pages.</li>
    </ol>
    <p style="margin-top:8px;color:var(--muted);font-size:12px">Если настроен GitHub Actions — достаточно загрузить data_entry.xlsx через интерфейс GitHub, HTML перегенерируется автоматически.</p>
  </div>

</div>
</div><!-- /app-content -->

<script>
const PW_HASH = 'fa7496e4ae840306df41bd658800392e24db8ac4767159d6bdf8a23b28c44ea0';
async function sha256(s) {{
  const buf = await crypto.subtle.digest('SHA-256', new TextEncoder().encode(s));
  return [...new Uint8Array(buf)].map(b => b.toString(16).padStart(2,'0')).join('');
}}
async function checkAuth() {{
  const pw = document.getElementById('auth-pw').value;
  const hash = await sha256(pw);
  if (hash === PW_HASH) {{
    sessionStorage.setItem('auth','1');
    document.getElementById('auth-overlay').style.display='none';
    document.getElementById('app-content').style.display='block';
  }} else {{
    document.getElementById('auth-err').style.display='block';
  }}
}}
if (sessionStorage.getItem('auth')==='1') {{
  document.getElementById('auth-overlay').style.display='none';
  document.getElementById('app-content').style.display='block';
}}
</script>

<script>
const D = {json_data};
const STAGES = D.stages;
const EPOCH = new Date(2026, 0, 1);

// ── Dropdown Filters ──
function buildDD(listId, items, sortFn) {{
  const list = document.getElementById(listId);
  if (sortFn) items = [...items].sort(sortFn);
  items.forEach(v => {{
    const lbl = document.createElement('label');
    const cb = document.createElement('input');
    cb.type = 'checkbox'; cb.value = v; cb.checked = false;
    cb.onchange = () => {{ updateDDBtn(listId); render(); }};
    lbl.appendChild(cb); lbl.appendChild(document.createTextNode(v));
    list.appendChild(lbl);
  }});
}}

function toggleDD(ddId) {{
  const list = document.getElementById(ddId).querySelector('.dd-list');
  document.querySelectorAll('.dd-list.open').forEach(l => {{ if (l !== list) l.classList.remove('open'); }});
  list.classList.toggle('open');
}}

document.addEventListener('click', e => {{
  if (!e.target.closest('.dd')) document.querySelectorAll('.dd-list.open').forEach(l => l.classList.remove('open'));
}});

function getDDValues(listId) {{
  return [...document.getElementById(listId).querySelectorAll('input:checked')].map(cb => cb.value);
}}

function updateDDBtn(listId) {{
  const dd = document.getElementById(listId).closest('.dd');
  const btn = dd.querySelector('.dd-btn');
  const checked = getDDValues(listId);
  if (checked.length === 0) {{
    btn.textContent = 'Все'; btn.classList.remove('has-selection');
  }} else if (checked.length <= 2) {{
    btn.textContent = checked.join(', '); btn.classList.add('has-selection');
  }} else {{
    btn.textContent = checked.length + ' выбрано'; btn.classList.add('has-selection');
  }}
}}

function initFilters() {{
  buildDD('ddl-prod', D.products, (a,b) => prodOrd(a) - prodOrd(b));
  buildDD('ddl-cat', D.categories);
  buildDD('ddl-instr', D.instruments);
}}

function matchFilter(vals, v) {{
  return vals.length === 0 || vals.includes(v);
}}

function clearFilters() {{
  document.querySelectorAll('.dd-list input[type=checkbox]').forEach(cb => cb.checked = false);
  document.querySelectorAll('.dd-btn').forEach(b => {{ b.textContent = 'Все'; b.classList.remove('has-selection'); }});
  document.getElementById('f-hide-inactive').checked = false;
  render();
}}

function getFiltered() {{
  const fp = getDDValues('ddl-prod');
  const fc = getDDValues('ddl-cat');
  const fi = getDDValues('ddl-instr');
  return D.rows.filter(r => {{
    if (!r.active) return false;
    if (!matchFilter(fp, r.prod)) return false;
    if (!matchFilter(fc, r.cat)) return false;
    if (!matchFilter(fi, r.instr)) return false;
    return true;
  }});
}}

// ── Tabs ──
function switchTab(t) {{
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.toggle('active', b.dataset.tab === t));
  document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
  document.getElementById('tab-' + t).classList.add('active');
  render();
}}

// ── Product & category sort ──
const PROD_ORDER = {{
  'КК':0,'КН':1,'КЛ':2,'КНР':3,'КНО':4,'POS':5,'BNPL':6,'Долями+':7,'Кубышка':8,
  'Незалоги.Дабл':9,
  'Авто':11,'Недвижимость':12,'Залоги.Дабл':13,
  'Умершие':20,'Банкроты':21,'Нерезиденты':22,'3P':23,'Инсталлмент':24
}};
const CAT_ORDER = {{'PRE':0,'1':1,'2':2,'3':3,'4':4}};
function prodOrd(p) {{ return PROD_ORDER[p] ?? 50; }}
function catOrd(c) {{ return CAT_ORDER[c] ?? 9; }}
function cmpProdCat(pA,cA,pB,cB) {{
  const dp = prodOrd(pA) - prodOrd(pB);
  return dp !== 0 ? dp : catOrd(cA) - catOrd(cB);
}}

// ── Helpers ──
function fmtPct(v) {{ return (v * 100).toFixed(1) + '%'; }}
function fmtDate(s) {{
  if (!s) return '';
  const d = new Date(s);
  return String(d.getDate()).padStart(2,'0') + '.' + String(d.getMonth()+1).padStart(2,'0') + '.' + String(d.getFullYear()).slice(2);
}}
function ragClass(r) {{ return r && r !== '—' ? 'rag-' + r : ''; }}
function progBar(v) {{
  return `<div class="prog-bar"><div class="prog-fill" style="width:${{Math.round(v*100)}}%"></div></div> ${{fmtPct(v)}}`;
}}
function dateToDays(s) {{
  if (!s) return null;
  return Math.round((new Date(s) - EPOCH) / 86400000);
}}

// ── KPIs ──
function renderKPIs() {{
  const rows = getFiltered();
  const tw = rows.reduce((s, r) => s + r.weight, 0) || 1;
  const prog = rows.reduce((s, r) => s + r.weight * r.progress, 0) / tw;
  const red = rows.filter(r => r.rag === 'RED').length;
  const done = rows.filter(r => r.rag === 'DONE').length;
  document.getElementById('kpis').innerHTML = `
    <div class="kpi"><div class="val">${{fmtPct(prog)}}</div><div class="lbl">Прогресс <span class="info-tip" title="Средневзвешенный прогресс по весам продуктов. Вес этапов: Старт 5%, ИТ 10%, 1%→20%, 5%→40%, 50%→75%, 100%→100%">ⓘ</span></div></div>
    <div class="kpi"><div class="val">${{rows.length}}</div><div class="lbl">Активных <span class="info-tip" title="Количество строк с Активен=Да, прошедших через текущие фильтры">ⓘ</span></div></div>
    <div class="kpi red"><div class="val">${{red}}</div><div class="lbl">RED <span class="info-tip" title="Ближайший плановый этап просрочен более чем на 14 дней (план vs сегодня)">ⓘ</span></div></div>
    <div class="kpi green"><div class="val">${{done}}</div><div class="lbl">DONE <span class="info-tip" title="Все 6 этапов завершены — фактическая дата 100% заполнена">ⓘ</span></div></div>
    <div class="kpi"><div class="val">${{rows.filter(r => r.rag === 'AMBER').length}}</div><div class="lbl">AMBER <span class="info-tip" title="Ближайший плановый этап просрочен на 1-14 дней">ⓘ</span></div></div>
    <div class="kpi"><div class="val">${{rows.filter(r => r.rag === 'GREEN').length}}</div><div class="lbl">GREEN <span class="info-tip" title="Ближайший плановый этап ещё не наступил — всё в плане">ⓘ</span></div></div>`;
}}

// ── Dashboard matrix ──
function renderDashboard() {{
  const fp = getDDValues('ddl-prod');
  const fc = getDDValues('ddl-cat');
  const fi = getDDValues('ddl-instr');
  const allRows = D.rows.filter(r => {{
    if (!matchFilter(fp, r.prod)) return false;
    if (!matchFilter(fc, r.cat)) return false;
    if (!matchFilter(fi, r.instr)) return false;
    return true;
  }});
  const hideInactive = document.getElementById('f-hide-inactive').checked;
  let segs = [...new Set(allRows.map(r => r.seg))].sort((a,b) => {{
    const [pA,cA] = a.split(' | '); const [pB,cB] = b.split(' | ');
    return cmpProdCat(pA,cA,pB,cB);
  }});
  const instrs = [...new Set(allRows.map(r => r.instr))].sort();
  if (hideInactive) {{
    segs = segs.filter(seg => allRows.some(r => r.seg === seg && r.active));
  }}
  if (!segs.length) {{ document.getElementById('dash-body').innerHTML = '<p>Нет данных</p>'; return; }}
  let h = '';
  const PROD_GRP = {{
    'КК':'Незалоговые','КН':'Незалоговые','КЛ':'Незалоговые','КНР':'Незалоговые','КНО':'Незалоговые',
    'POS':'Незалоговые','BNPL':'Незалоговые','Долями+':'Незалоговые','Кубышка':'Незалоговые','Незалоги.Дабл':'Незалоговые',
    'Авто':'Залоговые','Недвижимость':'Залоговые','КЛ':'Незалоговые','Залоги.Дабл':'Залоговые',
    'Умершие':'Спецсегменты','Банкроты':'Спецсегменты','Нерезиденты':'Спецсегменты','3P':'Спецсегменты','Инсталлмент':'Спецсегменты'
  }};
  const nCols = instrs.length + 1;
  h += '<div class="matrix-wrap"><table class="matrix"><thead><tr><th>Сегмент</th>';
  instrs.forEach(i => h += `<th class="instr">${{i}}</th>`);
  h += '</tr></thead><tbody>';
  let lastGrp = '';
  segs.forEach(seg => {{
    const prod = seg.split(' | ')[0];
    const grp = PROD_GRP[prod] || 'Прочие';
    if (grp !== lastGrp) {{
      h += `<tr><td colspan="${{nCols}}" style="background:var(--blue);color:#fff;font-weight:700;font-size:13px;padding:6px 10px;text-align:left">${{grp}}</td></tr>`;
      lastGrp = grp;
    }}
    h += `<tr><td class="left" style="font-weight:600">${{seg}}</td>`;
    instrs.forEach(instr => {{
      const all = allRows.filter(r => r.seg === seg && r.instr === instr);
      const active = all.filter(r => r.active);
      if (!active.length) {{
        // Inactive: show as grayed out, not counted in progress
        h += `<td style="background:#F0F0F0;color:#ccc;font-size:10px">—</td>`;
      }} else {{
        const tw = active.reduce((s, r) => s + r.weight, 0) || 1;
        const p = active.reduce((s, r) => s + r.weight * r.progress, 0) / tw;
        const alpha = 0.15 + p * 0.65;
        const bg = `rgba(47,84,150,${{alpha.toFixed(2)}})`;
        const fg = p > 0.4 ? '#fff' : '#333';
        h += `<td style="background:${{bg}};color:${{fg}}">${{p > 0 ? fmtPct(p) : '0%'}}</td>`;
      }}
    }});
    h += '</tr>';
  }});
  h += '</tbody></table></div>';
  document.getElementById('dash-body').innerHTML = h;
}}

// ── Timeline ──
function renderTimeline() {{
  const rows = getFiltered();
  const hideInactive = document.getElementById('f-hide-inactive').checked;
  const detailed = document.getElementById('f-tl-detail').checked;
  let h = '';

  if (detailed) {{
    // ── Detailed: one row per subproduct × instrument ──
    let detail = rows.map(r => ({{...r}}));
    if (hideInactive) detail = detail.filter(r => r.plans.some(Boolean) || r.facts.some(Boolean));
    detail.sort((a,b) => {{
      const pc = cmpProdCat(a.prod, a.cat, b.prod, b.cat);
      if (pc !== 0) return pc;
      if (a.subprod !== b.subprod) return a.subprod.localeCompare(b.subprod);
      return a.instr.localeCompare(b.instr);
    }});

    h += '<div style="overflow-x:auto"><table><thead><tr><th>Сегмент</th><th>Подпродукт</th><th>Подсегмент</th><th>Инструмент</th>';
    STAGES.forEach(s => h += `<th colspan="2" class="stage-grp">${{s}}</th>`);
    h += '<th>Прогресс</th><th>RAG</th></tr><tr><th></th><th></th><th></th><th></th>';
    STAGES.forEach(() => h += '<th style="font-size:10px">план</th><th style="font-size:10px">факт</th>');
    h += '<th></th><th></th></tr></thead><tbody>';

    detail.forEach(r => {{
      h += `<tr><td class="left" style="font-weight:600">${{r.seg}}</td>`;
      h += `<td class="left">${{r.subprod}}</td><td>${{r.subseg||''}}</td><td>${{r.instr}}</td>`;
      for (let i = 0; i < STAGES.length; i++) {{
        const pc = r.plans[i] ? 'plan-cell' : '';
        const fc = r.facts[i] ? 'fact-done' : '';
        h += `<td class="${{pc}}">${{fmtDate(r.plans[i])}}</td><td class="${{fc}}">${{fmtDate(r.facts[i])}}</td>`;
      }}
      h += `<td>${{progBar(r.progress)}}</td><td class="${{ragClass(r.rag)}}">${{r.rag}}</td></tr>`;
    }});
    h += '</tbody></table></div>';

  }} else {{
    // ── Aggregated: segment × instrument ──
    const groups = {{}};
    rows.forEach(r => {{
      const k = r.seg + '|' + r.instr;
      if (!groups[k]) groups[k] = {{ seg: r.seg, prod: r.prod, cat: r.cat, instr: r.instr, items: [] }};
      groups[k].items.push(r);
    }});
    let entries = Object.values(groups).sort((a,b) => {{
      const pc = cmpProdCat(a.prod, a.cat, b.prod, b.cat);
      return pc !== 0 ? pc : a.instr.localeCompare(b.instr);
    }});
    if (hideInactive) {{
      entries = entries.filter(g => g.items.some(r => r.plans.some(Boolean) || r.facts.some(Boolean)));
    }}

    h += '<table><thead><tr><th>Сегмент</th><th>Инструмент</th>';
    STAGES.forEach(s => h += `<th colspan="2" class="stage-grp">${{s}}</th>`);
    h += '<th>Прогресс</th><th>RAG</th></tr><tr><th></th><th></th>';
    STAGES.forEach(() => h += '<th style="font-size:10px">план</th><th style="font-size:10px">факт</th>');
    h += '<th></th><th></th></tr></thead><tbody>';

    entries.forEach(g => {{
      const tw = g.items.reduce((s,r) => s + r.weight, 0) || 1;
      const prog = g.items.reduce((s,r) => s + r.weight * r.progress, 0) / tw;
      const plans = STAGES.map((_,i) => {{
        const dates = g.items.map(r => r.plans[i]).filter(Boolean);
        return dates.length ? dates.sort()[0] : null;
      }});
      const facts = STAGES.map((_,i) => {{
        const dates = g.items.map(r => r.facts[i]).filter(Boolean);
        return dates.length ? dates.sort().pop() : null;
      }});
      const rags = g.items.map(r => r.rag);
      const rag = rags.includes('RED') ? 'RED' : rags.includes('AMBER') ? 'AMBER' :
                  rags.every(r => r === 'DONE') && rags.length ? 'DONE' : 'GREEN';

      h += `<tr><td class="left" style="font-weight:600">${{g.seg}}</td><td>${{g.instr}}</td>`;
      STAGES.forEach((_,i) => {{
        const pc = plans[i] ? 'plan-cell' : '';
        const fc = facts[i] ? 'fact-done' : '';
        h += `<td class="${{pc}}">${{fmtDate(plans[i])}}</td><td class="${{fc}}">${{fmtDate(facts[i])}}</td>`;
      }});
      h += `<td>${{progBar(prog)}}</td><td class="${{ragClass(rag)}}">${{rag}}</td></tr>`;
    }});
    h += '</tbody></table>';
  }}
  document.getElementById('tl-body').innerHTML = h;
}}

// ── Gantt ──
const STAGE_COLORS = ['#264653','#2A9D8F','#E9C46A','#F4A261','#E76F51','#E63946'];
const STAGE_COLORS_LIGHT = ['#26465366','#2A9D8F66','#E9C46A66','#F4A26166','#E76F5166','#E6394666'];

function ganttStageDurs(dates, isFact) {{
  // Compute per-stage durations from an array of day-numbers
  // isFact: if true, last completed stage extends to today (ongoing)
  const todayD = Math.round((new Date() - EPOCH) / 86400000);
  const durs = [];
  for (let i = 0; i < STAGES.length; i++) {{
    let dur = 0;
    if (dates[i] !== null) {{
      let next = null;
      for (let j = i + 1; j < STAGES.length; j++) {{
        if (dates[j] !== null) {{ next = dates[j]; break; }}
      }}
      if (next !== null) {{
        dur = next - dates[i];
      }} else if (isFact) {{
        // Last completed stage — extends to today
        dur = Math.max(todayD - dates[i], 1);
      }} else {{
        dur = 30; // plan: default 30 days for last stage
      }}
      if (dur < 0) dur = 0;
    }}
    durs.push(dur);
  }}
  return durs;
}}

function renderGantt() {{
  const rows = getFiltered();
  const container = document.getElementById('gantt-body');
  const mode = document.getElementById('gantt-mode').value;
  const detail = document.getElementById('gantt-detail').checked;
  container.innerHTML = '';

  D.categories.forEach(cat => {{
    const catRows = rows.filter(r => r.cat === cat);
    if (!catRows.length) return;

    // Group by Product × Instrument (optionally with subproduct detail)
    const groups = {{}};
    catRows.forEach(r => {{
      let key = r.prod + ' | ' + r.instr;
      if (detail) {{
        key = r.prod + (r.subprod ? ' / ' + r.subprod : '') + (r.subseg ? ' (' + r.subseg + ')' : '') + ' | ' + r.instr;
      }}
      if (!groups[key]) groups[key] = {{ plans: STAGES.map(()=>null), facts: STAGES.map(()=>null), _prod: r.prod, _instr: r.instr }};
      const g = groups[key];
      for (let i = 0; i < STAGES.length; i++) {{
        const pd = dateToDays(r.plans[i]);
        const fd = dateToDays(r.facts[i]);
        if (pd !== null && (g.plans[i] === null || pd < g.plans[i])) g.plans[i] = pd;
        if (fd !== null && (g.facts[i] === null || fd > g.facts[i])) g.facts[i] = fd;
      }}
    }});

    const sorted = Object.entries(groups).sort((a,b) => {{
      const dp = prodOrd(a[1]._prod) - prodOrd(b[1]._prod);
      return dp !== 0 ? dp : a[1]._instr.localeCompare(b[1]._instr);
    }}).filter(([_,d]) => !d.plans.every(v => v === null));
    if (!sorted.length) return;

    // Build data based on mode
    const labels = [], allTraces = [];
    const showPlan = mode === 'plan' || mode === 'compare';
    const showFact = mode === 'fact' || mode === 'compare';

    if (mode === 'compare') {{
      // Two rows per product: "Prod (план)" and "Prod (факт)"
      const gapPlan = [], gapFact = [];
      const planDurs = STAGES.map(() => []);
      const factDurs = STAGES.map(() => []);

      sorted.forEach(([prod, d]) => {{
        labels.push(prod + ' (план)');
        labels.push(prod + ' (факт)');
        const fp = d.plans.find(v => v !== null) || 0;
        const ff = d.facts.find(v => v !== null) || fp;
        gapPlan.push(fp); gapFact.push(ff);
        const pd = ganttStageDurs(d.plans, false);
        const fd = ganttStageDurs(d.facts, true);
        for (let i = 0; i < STAGES.length; i++) {{
          planDurs[i].push(pd[i]); planDurs[i].push(0);
          factDurs[i].push(0); factDurs[i].push(fd[i]);
        }}
      }});

      // Interleave gaps
      const gapAll = [];
      for (let j = 0; j < sorted.length; j++) {{ gapAll.push(gapPlan[j]); gapAll.push(gapFact[j]); }}

      allTraces.push({{ type:'bar', orientation:'h', y:labels, x:gapAll, name:'', showlegend:false,
        marker:{{color:'rgba(0,0,0,0)'}}, hoverinfo:'skip' }});
      for (let i = 0; i < STAGES.length; i++) {{
        // Plan bars: lighter color
        const planX = planDurs[i];
        const factX = factDurs[i];
        const combined = planX.map((p,j) => p + factX[j]);
        const colors = combined.map((_, j) => j % 2 === 0 ? STAGE_COLORS_LIGHT[i] : STAGE_COLORS[i]);
        allTraces.push({{
          type:'bar', orientation:'h', y:labels, x:combined, name: STAGES[i],
          marker:{{ color: colors }}, showlegend: i === 0 ? false : true,
          legendgroup: STAGES[i],
          hovertemplate: STAGES[i] + ': %{{x}} дн.<extra></extra>'
        }});
      }}

    }} else {{
      // Single mode: plan or fact
      const useDates = showPlan ? 'plans' : 'facts';
      const gapArr = [];
      const durs = STAGES.map(() => []);

      sorted.forEach(([prod, d]) => {{
        const dates = d[useDates];
        if (dates.every(v => v === null)) {{ return; }}
        labels.push(prod);
        const first = dates.find(v => v !== null) || 0;
        gapArr.push(first);
        const sd = ganttStageDurs(dates, showFact);
        for (let i = 0; i < STAGES.length; i++) durs[i].push(sd[i]);
      }});

      if (!labels.length) return;

      allTraces.push({{ type:'bar', orientation:'h', y:labels, x:gapArr, name:'', showlegend:false,
        marker:{{color:'rgba(0,0,0,0)'}}, hoverinfo:'skip' }});
      for (let i = 0; i < STAGES.length; i++) {{
        allTraces.push({{
          type:'bar', orientation:'h', y:labels, x:durs[i], name: STAGES[i],
          marker:{{ color: showFact ? STAGE_COLORS[i] : STAGE_COLORS_LIGHT[i] }},
          hovertemplate: STAGES[i] + ': %{{x}} дн.<extra></extra>'
        }});
      }}
    }}

    if (!labels.length) return;

    const div = document.createElement('div');
    div.className = 'gantt-section';
    container.appendChild(div);

    const tv = [], tt = [];
    const mNames = ['Янв','Фев','Мар','Апр','Май','Июн','Июл','Авг','Сен','Окт','Ноя','Дек'];
    for (let y = 2026; y <= 2028; y++)
      for (let m = 0; m < 12; m++) {{
        const days = Math.round((new Date(y, m, 1) - EPOCH) / 86400000);
        if (days >= -30 && days <= 800) {{ tv.push(days); tt.push(mNames[m] + "'" + String(y).slice(2)); }}
      }}

    const todayDays = Math.round((new Date() - EPOCH) / 86400000);
    Plotly.newPlot(div, allTraces, {{
      barmode:'stack', title:'ПЗ ' + cat,
      xaxis:{{ tickvals:tv, ticktext:tt, gridcolor:'#eee',
               range:[0, Math.round((new Date(2026,11,31)-EPOCH)/864e5)] }},
      yaxis:{{ autorange:'reversed' }},
      height: Math.max(labels.length * 28 + 120, 250),
      margin:{{ l:220, r:20, t:40, b:40 }},
      showlegend: false,
      shapes: [{{
        type: 'line', x0: todayDays, x1: todayDays, y0: 0, y1: 1, yref: 'paper',
        line: {{ color: '#C00000', width: 2, dash: 'dash' }}
      }}],
      annotations: [{{
        x: todayDays, y: 1.02, yref: 'paper', text: 'Сегодня', showarrow: false,
        font: {{ size: 11, color: '#C00000' }}
      }}]
    }}, {{ responsive:true }});
  }});

  if (!container.querySelectorAll('.gantt-section').length) {{
    container.innerHTML = '<p>Нет данных для Ганта (нет плановых дат)</p>';
  }}
}}

// ── Data table ──
function renderData() {{
  const hideEmpty = document.getElementById('f-hide-inactive').checked;
  let rows = getFiltered();
  if (hideEmpty) {{
    rows = rows.filter(r => r.plans.some(Boolean) || r.facts.some(Boolean));
  }}
  let h = `<div style="overflow-x:auto;max-width:100%"><table style="min-width:800px"><thead><tr>
    <th>Продукт</th><th>Подпродукт</th><th>Кат.</th><th>Инструмент</th>
    <th>Этап</th><th>Прогресс</th><th>RAG</th><th>Сдвиг</th>`;
  STAGES.forEach(s => h += `<th>${{s}} п</th><th>${{s}} ф</th>`);
  h += `<th>Эпики</th><th>Комментарии</th>
  </tr></thead><tbody>`;
  rows.forEach(r => {{
    h += `<tr>
      <td class="left">${{r.prod}}</td><td class="left">${{r.subprod}}</td>
      <td>${{r.cat}}</td><td>${{r.instr}}</td>
      <td>${{r.stage}}</td><td>${{progBar(r.progress)}}</td>
      <td class="${{ragClass(r.rag)}}">${{r.rag}}</td>
      <td>${{r.slip !== null ? r.slip : ''}}</td>`;
    for (let i = 0; i < STAGES.length; i++) {{
      h += `<td class="plan-cell">${{fmtDate(r.plans[i])}}</td>`;
      h += `<td class="${{r.facts[i] ? 'fact-done' : ''}}">${{fmtDate(r.facts[i])}}</td>`;
    }}
    h += `<td class="left">${{r.epics||''}}</td><td class="left">${{r.comment||''}}</td></tr>`;
  }});
  h += '</tbody></table></div>';
  document.getElementById('data-body').innerHTML = h;
}}

// ── Render all ──
function render() {{
  renderKPIs();
  const active = document.querySelector('.tab-btn.active');
  const tab = active ? active.dataset.tab : 'dashboard';
  if (tab === 'dashboard') renderDashboard();
  else if (tab === 'timeline') renderTimeline();
  else if (tab === 'gantt') renderGantt();
  else if (tab === 'data') renderData();
  else if (tab === 'editor') edUpdatePreview();
}}

// ── Init tab controls (once, never re-rendered) ──
function initTabControls() {{
  const selStyle = 'padding:5px 10px;border:1px solid var(--border);border-radius:4px;font-size:13px';
  // Dashboard
  document.getElementById('dash-ctrl').innerHTML =
    '<div class="legend" style="margin:0"><span><i style="background:rgba(47,84,150,0.15)"></i> Активно (0%)</span>'
    + '<span><i style="background:rgba(47,84,150,0.7)"></i> Активно (прогресс)</span>'
    + '<span><i style="background:#F0F0F0;border:1px solid #ddd"></i> Не предусмотрено</span></div>';

  // Timeline
  document.getElementById('tl-ctrl').innerHTML =
    '<label><input type="checkbox" id="f-tl-detail" onchange="render()"> Детализация по подпродуктам</label>';

  // Gantt
  document.getElementById('gantt-ctrl').innerHTML =
    '<div><label style="font-size:12px;color:var(--muted)">Режим</label><br>'
    + `<select id="gantt-mode" onchange="render()" style="${{selStyle}}">`
    + '<option value="compare">Сравнение план/факт</option>'
    + '<option value="plan">Только план</option>'
    + '<option value="fact">Только факт</option></select></div>'
    + '<label><input type="checkbox" id="gantt-detail" onchange="render()"> Детализация по подпродуктам</label>'
    + '<div class="legend" style="margin:0">'
    + STAGES.map((s,i) => `<span><i style="background:${{STAGE_COLORS[i]}}"></i> ${{s}}</span>`).join('')
    + '</div>'
    + '<div class="legend" style="margin:0;border-left:1px solid var(--border);padding-left:12px">'
    + '<span><i style="background:#26465366;border:1px dashed #264653"></i> План</span>'
    + '<span><i style="background:#264653"></i> Факт</span>'
    + '<span style="color:var(--red);font-weight:600">┆ Сегодня</span>'
    + '</div>';

  // Data — no extra controls

  // Editor — full inline editing UI
  initEditor();
}}

// ── Editor ──
function initEditor() {{
  const ec = document.getElementById('editor-ctrl');
  ec.innerHTML = '<span style="font-size:13px;color:var(--muted)">Массовое редактирование: выберите строки → установите значения → Применить → Скачать XLSX</span>';

  const body = document.getElementById('editor-body');
  const stageOpts = STAGES.map(s => `<option value="${{s}}">${{s}}</option>`).join('');

  // Build unique subproducts list
  const subprods = [...new Set(D.rows.map(r => r.subprod).filter(Boolean))].sort();

  body.innerHTML = `
  <div class="ed-section">
    <h4>1. Выберите строки</h4>
    <div class="ed-filters">
      <div><label style="font-size:12px;color:var(--muted)">Продукт</label><br>
        <div class="dd" id="dd-ed-prod"><button class="dd-btn" onclick="toggleDD('dd-ed-prod')">Все</button>
        <div class="dd-list" id="ddl-ed-prod"></div></div></div>
      <div><label style="font-size:12px;color:var(--muted)">Подпродукт</label><br>
        <div class="dd" id="dd-ed-sub"><button class="dd-btn" onclick="toggleDD('dd-ed-sub')">Все</button>
        <div class="dd-list" id="ddl-ed-sub"></div></div></div>
      <div><label style="font-size:12px;color:var(--muted)">Категория ПЗ</label><br>
        <div class="dd" id="dd-ed-cat"><button class="dd-btn" onclick="toggleDD('dd-ed-cat')">Все</button>
        <div class="dd-list" id="ddl-ed-cat"></div></div></div>
      <div><label style="font-size:12px;color:var(--muted)">Инструмент</label><br>
        <div class="dd" id="dd-ed-instr"><button class="dd-btn" onclick="toggleDD('dd-ed-instr')">Все</button>
        <div class="dd-list" id="ddl-ed-instr"></div></div></div>
    </div>
    <div class="ed-count" id="ed-count">Выбрано: 0 строк</div>
  </div>

  <div class="ed-section">
    <h4>2. Установить Активен</h4>
    <div class="ed-actions">
      <select id="ed-active-val"><option value="">—</option><option value="Да">Да</option><option value="Нет">Нет</option></select>
      <button class="ed-btn ed-btn-primary" onclick="edApplyActive()">Применить Активен</button>
    </div>
  </div>

  <div class="ed-section">
    <h4>3. Установить даты этапа</h4>
    <div class="ed-actions">
      <div><label style="font-size:12px;color:var(--muted)">Этап</label><br>
        <select id="ed-stage">${{stageOpts}}</select></div>
      <div><label style="font-size:12px;color:var(--muted)">План</label><br>
        <input type="date" id="ed-plan"></div>
      <div><label style="font-size:12px;color:var(--muted)">Факт</label><br>
        <input type="date" id="ed-fact"></div>
      <button class="ed-btn ed-btn-primary" onclick="edApplyDates()">Применить даты</button>
    </div>
  </div>

  <div class="ed-section">
    <h4>4. Сохранить</h4>
    <div class="ed-actions">
      <button class="ed-btn ed-btn-success" onclick="edPushToGH()">🚀 Опубликовать в GitHub</button>
      <button class="ed-btn" style="background:var(--gray);color:var(--text)" onclick="edDownload()">⬇ Скачать XLSX</button>
      <span id="ed-push-status" style="font-size:12px"></span>
    </div>
    <div id="gh-settings" style="display:none;margin-top:8px;padding:8px;background:var(--gray);border-radius:6px">
      <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">
        <input type="password" id="gh-token" placeholder="GitHub Personal Access Token"
          style="padding:6px 10px;border:1px solid var(--border);border-radius:4px;font-size:12px;width:320px"
          value="">
        <button class="ed-btn" style="background:var(--blue);color:#fff;font-size:12px;padding:4px 12px"
          onclick="localStorage.setItem('gh_token',document.getElementById('gh-token').value);document.getElementById('gh-settings').style.display='none';alert('Токен сохранён')">Сохранить</button>
        <span style="font-size:11px;color:var(--muted)">Нужен токен с правами repo. <a href="https://github.com/settings/tokens/new?scopes=repo&description=Pereezd+Roadmap" target="_blank">Создать →</a></span>
      </div>
    </div>
  </div>

  <div class="ed-section">
    <h4>Предпросмотр выбранных строк</h4>
    <div class="ed-preview" id="ed-preview"></div>
  </div>`;

  // Initial build — will be refreshed dynamically
  edRefreshFilters();
}}

function edRefreshFilters() {{
  // Get current selections before rebuilding
  const prevProd = getDDValues('ddl-ed-prod');
  const prevSub = getDDValues('ddl-ed-sub');
  const prevCat = getDDValues('ddl-ed-cat');
  const prevInstr = getDDValues('ddl-ed-instr');

  // Filter rows progressively: each dropdown limits options for the others
  let pool = D.rows;
  if (prevProd.length) pool = pool.filter(r => prevProd.includes(r.prod));
  const availSub = [...new Set(pool.map(r => r.subprod).filter(Boolean))].sort();

  let pool2 = pool;
  if (prevSub.length) pool2 = pool2.filter(r => prevSub.includes(r.subprod));
  const availCat = [...new Set(pool2.map(r => r.cat))].sort((a,b) => {{
    const o = {{'PRE':0,'1':1,'2':2,'3':3,'4':4}}; return (o[a]??9) - (o[b]??9);
  }});

  let pool3 = pool2;
  if (prevCat.length) pool3 = pool3.filter(r => prevCat.includes(r.cat));
  const availInstr = [...new Set(pool3.map(r => r.instr))].sort();

  // Also compute available products based on other selections (reverse direction)
  let poolRev = D.rows;
  if (prevSub.length) poolRev = poolRev.filter(r => prevSub.includes(r.subprod));
  if (prevCat.length) poolRev = poolRev.filter(r => prevCat.includes(r.cat));
  if (prevInstr.length) poolRev = poolRev.filter(r => prevInstr.includes(r.instr));
  const availProd = [...new Set(poolRev.map(r => r.prod))].sort((a,b) => prodOrd(a) - prodOrd(b));

  // Rebuild each dropdown, preserving valid selections
  edRebuildDD('ddl-ed-prod', availProd, prevProd);
  edRebuildDD('ddl-ed-sub', availSub, prevSub);
  edRebuildDD('ddl-ed-cat', availCat, prevCat);
  edRebuildDD('ddl-ed-instr', availInstr, prevInstr);

  edUpdatePreview();
}}

function edRebuildDD(listId, items, selected) {{
  const list = document.getElementById(listId);
  list.innerHTML = '';
  items.forEach(v => {{
    const lbl = document.createElement('label');
    const cb = document.createElement('input');
    cb.type = 'checkbox'; cb.value = v;
    cb.checked = selected.includes(v);
    cb.onchange = () => {{ updateDDBtn(listId); edRefreshFilters(); }};
    lbl.appendChild(cb); lbl.appendChild(document.createTextNode(v));
    list.appendChild(lbl);
  }});
  updateDDBtn(listId);
}}

function edGetSelected() {{
  const fp = getDDValues('ddl-ed-prod');
  const fs = getDDValues('ddl-ed-sub');
  const fc = getDDValues('ddl-ed-cat');
  const fi = getDDValues('ddl-ed-instr');
  return D.rows.filter(r => {{
    if (fp.length && !fp.includes(r.prod)) return false;
    if (fs.length && !fs.includes(r.subprod)) return false;
    if (fc.length && !fc.includes(r.cat)) return false;
    if (fi.length && !fi.includes(r.instr)) return false;
    return true;
  }});
}}

function edUpdatePreview() {{
  const sel = edGetSelected();
  document.getElementById('ed-count').textContent = `Выбрано: ${{sel.length}} строк`;
  const preview = document.getElementById('ed-preview');
  if (!sel.length) {{ preview.innerHTML = '<p style="color:var(--muted)">Выберите фильтры выше</p>'; return; }}
  const show = sel.slice(0, 100);
  let h = '<table><thead><tr><th>Продукт</th><th>Подпродукт</th><th>Кат.</th><th>Инструмент</th><th>Активен</th>';
  STAGES.forEach(s => h += `<th>${{s}} п</th><th>${{s}} ф</th>`);
  h += '</tr></thead><tbody>';
  show.forEach(r => {{
    h += `<tr><td class="left">${{r.prod}}</td><td class="left">${{r.subprod}}</td><td>${{r.cat}}</td><td>${{r.instr}}</td>`;
    h += `<td style="font-weight:600;color:${{r.active?'var(--green)':'var(--muted)'}}">${{r.active?'Да':'Нет'}}</td>`;
    for (let i = 0; i < STAGES.length; i++) {{
      h += `<td class="plan-cell">${{fmtDate(r.plans[i])}}</td>`;
      h += `<td class="${{r.facts[i]?'fact-done':''}}">${{fmtDate(r.facts[i])}}</td>`;
    }}
    h += '</tr>';
  }});
  h += '</tbody></table>';
  if (sel.length > 100) h += `<p style="color:var(--muted)">...и ещё ${{sel.length - 100}} строк</p>`;
  preview.innerHTML = h;
}}

function edApplyActive() {{
  const val = document.getElementById('ed-active-val').value;
  if (!val) {{ alert('Выберите Да или Нет'); return; }}
  const sel = edGetSelected();
  if (!sel.length) {{ alert('Выберите строки'); return; }}
  const isActive = val === 'Да';
  sel.forEach(r => r.active = isActive);
  edUpdatePreview();
  render();
  alert(`Активен = ${{val}} установлен для ${{sel.length}} строк`);
}}

function edApplyDates() {{
  const stageIdx = STAGES.indexOf(document.getElementById('ed-stage').value);
  const planVal = document.getElementById('ed-plan').value || null;
  const factVal = document.getElementById('ed-fact').value || null;
  if (stageIdx < 0) {{ alert('Выберите этап'); return; }}
  if (!planVal && !factVal) {{ alert('Введите хотя бы одну дату'); return; }}
  const sel = edGetSelected();
  if (!sel.length) {{ alert('Выберите строки'); return; }}
  sel.forEach(r => {{
    if (planVal) r.plans[stageIdx] = planVal;
    if (factVal) r.facts[stageIdx] = factVal;
  }});
  edUpdatePreview();
  render();
  alert(`Даты этапа "${{STAGES[stageIdx]}}" обновлены для ${{sel.length}} строк`);
}}

function edBuildXLSX() {{
  const headers = {json.dumps(CSV_HEADERS, ensure_ascii=False)};
  const wsData = [headers];
  D.rows.forEach(r => {{
    const row = [r.agg, r.prod, r.subprod, r.subseg||'', r.cat, r.igrp, r.instr, r.active ? 'Да' : 'Нет'];
    STAGES.forEach((_, i) => row.push(r.plans[i] || ''));
    STAGES.forEach((_, i) => row.push(r.facts[i] || ''));
    STAGES.forEach(() => row.push(''));  // baseline
    row.push(r.epics || '', r.comment || '');
    wsData.push(row);
  }});
  const ws = XLSX.utils.aoa_to_sheet(wsData);
  const wb = XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb, ws, 'DATA');
  return wb;
}}

function edDownload() {{
  XLSX.writeFile(edBuildXLSX(), 'data_entry.xlsx');
}}

const GH_REPO = 'Annushkaev/pereezd-roadmap';
const GH_FILE = 'data_entry.xlsx';

async function edPushToGH() {{
  const token = localStorage.getItem('gh_token') || document.getElementById('gh-token').value;
  if (!token) {{
    document.getElementById('gh-settings').style.display = 'block';
    alert('Введите GitHub Token и нажмите Сохранить, затем попробуйте снова');
    return;
  }}
  const status = document.getElementById('ed-push-status');
  status.textContent = '⏳ Публикация...'; status.style.color = 'var(--blue)';

  try {{
    // 1. Generate XLSX as base64
    const wbout = XLSX.write(edBuildXLSX(), {{ type: 'base64', bookType: 'xlsx' }});

    // 2. Get current file SHA
    const getResp = await fetch(`https://api.github.com/repos/${{GH_REPO}}/contents/${{GH_FILE}}`, {{
      headers: {{ 'Authorization': `token ${{token}}` }}
    }});
    let sha = '';
    if (getResp.ok) {{
      const data = await getResp.json();
      sha = data.sha;
    }}

    // 3. Push file
    const putResp = await fetch(`https://api.github.com/repos/${{GH_REPO}}/contents/${{GH_FILE}}`, {{
      method: 'PUT',
      headers: {{
        'Authorization': `token ${{token}}`,
        'Content-Type': 'application/json'
      }},
      body: JSON.stringify({{
        message: 'Обновление данных из дашборда',
        content: wbout,
        sha: sha || undefined
      }})
    }});

    if (putResp.ok) {{
      status.textContent = '✅ Опубликовано! GitHub Actions перегенерирует дашборд через ~1 мин.';
      status.style.color = 'var(--green)';
    }} else {{
      const err = await putResp.json();
      throw new Error(err.message || putResp.status);
    }}
  }} catch(e) {{
    status.textContent = '❌ Ошибка: ' + e.message;
    status.style.color = 'var(--red)';
  }}
}}

// Load saved token + hidden shortcut Ctrl+Shift+G to show settings
document.addEventListener('DOMContentLoaded', () => {{
  const saved = localStorage.getItem('gh_token');
  if (saved) document.getElementById('gh-token').value = saved;
}});
document.addEventListener('keydown', e => {{
  if (e.ctrlKey && e.shiftKey && e.key === 'G') {{
    const s = document.getElementById('gh-settings');
    s.style.display = s.style.display === 'none' ? 'block' : 'none';
  }}
}});

initFilters();
initTabControls();
render();
</script>
</body>
</html>"""

    HTML_PATH.write_text(html, encoding="utf-8")
    print(f"  HTML: {HTML_PATH.name} ({HTML_PATH.stat().st_size // 1024} KB)")


# ── Main ─────────────────────────────────────────────────────────────

def main():
    force_csv = "--init" in sys.argv

    print("Parsing Confluence exports...")
    products = parse_products(ROADMAP_DIR / "Продукты+для+переезда.doc")
    instruments = parse_instruments(ROADMAP_DIR / "Инструменты+для+переезда.doc")
    print(f"  Products: {len(products)}, Instruments: {len(instruments)}")

    if force_csv or not ENTRY_PATH.exists():
        print("Generating entry XLSX...")
        generate_entry_xlsx(products, instruments)
    else:
        print(f"  XLSX: reading {ENTRY_PATH.name}")

    rows = read_entry_xlsx()
    print(f"  Rows: {len(rows)}")

    print("Computing derived values...")
    data = compute(rows, products)

    print("Generating HTML dashboard...")
    generate_html(data)
    print("Done!")

if __name__ == "__main__":
    main()
